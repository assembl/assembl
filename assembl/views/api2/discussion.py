# -*- coding: utf-8 -*-
from cStringIO import StringIO
from collections import defaultdict
from datetime import timedelta, datetime
from graphene.relay import Node

import isodate
import simplejson as json
import transaction
from cornice import Service
from pyramid.httpexceptions import (
    HTTPOk, HTTPBadRequest, HTTPUnauthorized, HTTPNotAcceptable, HTTPServerError, HTTPConflict)
from pyramid.renderers import JSONP_VALID_CALLBACK
from pyramid.response import Response
from pyramid.security import Everyone
from pyramid.settings import asbool
from pyramid.view import view_config
from pyramid_mailer import get_mailer
from pyramid_mailer.message import Message
from sqlalchemy import (
    Column,
    Integer,
    DateTime,
    cast,
    func,
    distinct,
    Table,
    MetaData,
    and_,
    or_,
    case,
    Float,
)
from sqlalchemy.orm import with_polymorphic, joinedload, joinedload_all
from sqlalchemy.orm.util import aliased
from sqlalchemy.sql.expression import literal

from assembl.auth import (
    P_READ, P_ADMIN_DISC, P_DISC_STATS, P_SYSADMIN,
    R_ADMINISTRATOR)
from assembl.auth.util import get_permissions, discussions_with_access
from assembl.graphql.utils import get_primary_id
from assembl.lib.clean_input import sanitize_text
from assembl.lib.config import get_config
from assembl.lib.json import DateJSONEncoder
from assembl.lib.migration import create_default_discussion_data
from assembl.lib.parsedatetime import parse_datetime
from assembl.lib.sqla import ObjectNotUniqueError
from assembl.models import Discussion, Post
from assembl.models import LanguagePreferenceCollection, Locale, PublicationStates
from assembl.models.idea import MessageView
from assembl.models.idea_content_link import ExtractStates
from assembl.models.post import deleted_publication_states
from assembl.models.social_data_extraction import (
    get_social_columns_from_user, load_social_columns_info, get_provider_id_for_discussion)
from assembl.utils import (
    format_date,
    get_thread_ideas, get_survey_ideas, get_multicolumns_ideas,
    get_bright_mirror_ideas, get_vote_session_ideas,
    get_deleted_posts, get_related_extracts, get_posts,
    get_published_posts, get_published_top_posts)
from . import (JSON_HEADER, FORM_HEADER, CreationResponse)
from ..api.discussion import etalab_discussions, API_ETALAB_DISCUSSIONS_PREFIX
from ..traversal import InstanceContext, ClassContext

no_thematic_associated = "no thematic associated"

sheet_names = ["export_phase",
               "export_module_survey",
               "export_module_thread",
               "export_module_multicolumns",
               "export_module_vote",
               "vote_users_data",
               "export_module_bright_mirror"]

@view_config(context=InstanceContext, request_method='GET',
             ctx_instance_class=Discussion, permission=P_READ,
             accept="application/json", name="settings",
             renderer='json')
def discussion_settings_get(request):
    return request.context._instance.settings_json


@view_config(context=InstanceContext, request_method='PATCH',
             ctx_instance_class=Discussion, permission=P_ADMIN_DISC,
             header=JSON_HEADER, name="settings")
@view_config(context=InstanceContext, request_method='PUT',
             ctx_instance_class=Discussion, permission=P_ADMIN_DISC,
             header=JSON_HEADER, name="settings")
def discussion_settings_put(request):
    request.context._instance.settings_json = request.json_body
    return HTTPOk()


def handle_jsonp(callback_fn, json):
    # TODO: Use an augmented JSONP renderer with ld content-type
    if not JSONP_VALID_CALLBACK.match(callback_fn):
        raise HTTPBadRequest("invalid callback name")
    return "/**/{0}({1});".format(callback_fn.encode('ascii'), json)


JSON_MIMETYPE = 'application/json'
CSV_MIMETYPE = 'text/csv'
XSL_MIMETYPE = 'application/vnd.ms-excel'
XSLX_MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

stats_formats_mimetypes = {
    'json': JSON_MIMETYPE,
    'csv': CSV_MIMETYPE,
    'xlsx': XSLX_MIMETYPE,
    'xls': XSL_MIMETYPE,
}

# ordered by preference
default_stats_formats = [XSLX_MIMETYPE, JSON_MIMETYPE, CSV_MIMETYPE]


def get_format(request, stats_formats=default_stats_formats):
    format = request.GET.get('format', None)
    if format:
        format = stats_formats_mimetypes.get(format, None)
        if not format:
            raise HTTPBadRequest("format: use one of " + ", ".join(
                [k for (k, v) in stats_formats_mimetypes.iteritems()
                 if v in stats_formats]))
    else:
        format = request.accept.best_match(stats_formats)
        if not format:
            raise HTTPNotAcceptable("Use one of " + ", ".join(stats_formats))
    return format


def get_time_series_timing(request, force_bounds=False):
    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    interval = request.GET.get("interval", None)
    try:
        if start:
            start = parse_datetime(start)
            if force_bounds and start:
                discussion = request.context._instance
                discussion_lower_bound = discussion.creation_date
                if start < discussion_lower_bound:
                    start = discussion_lower_bound
        else:
            discussion = request.context._instance
            start = discussion.creation_date
            # TODO: Round down at day/week/month according to interval
        if end:
            end = parse_datetime(end)
            if force_bounds and end:
                if end < start:
                    end = start
                discussion_upper_bound = datetime.now()
                if end > discussion_upper_bound:
                    end = discussion_upper_bound
        else:
            end = datetime.now()
        if interval:
            interval = isodate.parse_duration(interval)
        else:
            interval = end - start + timedelta(seconds=1)
    except isodate.ISO8601Error as e:
        raise HTTPBadRequest(e)
    return (start, end, interval)


@view_config(context=InstanceContext, name="time_series_analytics",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_DISC_STATS)
def get_time_series_analytics(request):
    start, end, interval = get_time_series_timing(request)
    discussion = request.context._instance
    user_id = request.authenticated_userid or Everyone
    format = get_format(request)
    results = []

    with transaction.manager:
        bind = discussion.db.connection()
        metadata = MetaData(discussion.db.get_bind())  # make sure we are using the same connexion

        intervals_table = Table('temp_table_intervals_' + str(user_id), metadata,
                                Column('interval_id', Integer, primary_key=True),
                                Column('interval_start', DateTime, nullable=False),
                                Column('interval_end', DateTime, nullable=False),
                                prefixes=['TEMPORARY']
                                )
        intervals_table.drop(bind=bind, checkfirst=True)
        intervals_table.create(bind=bind)
        interval_start = start
        intervals = []
        while interval_start < end:
            interval_end = min(interval_start + interval, end)
            intervals.append({'interval_start': interval_start, 'interval_end': interval_end})
            interval_start = interval_start + interval
        # pprint.pprint(intervals)
        discussion.db.execute(intervals_table.insert(), intervals)

        from assembl.models import (
            Post, AgentStatusInDiscussion, ViewPost, Idea,
            AbstractIdeaVote, ActionOnPost, ActionOnIdea, Content)

        # The posters
        post_subquery = discussion.db.query(intervals_table.c.interval_id,
                                            func.count(distinct(Post.id)).label('count_posts'),
                                            func.count(distinct(Post.creator_id)).label('count_post_authors'),
                                            # func.DB.DBA.BAG_AGG(Post.creator_id).label('post_authors'),
                                            # func.DB.DBA.BAG_AGG(Post.id).label('post_ids'),
                                            )
        post_subquery = post_subquery.outerjoin(Post, and_(
            Post.creation_date >= intervals_table.c.interval_start,
            Post.creation_date < intervals_table.c.interval_end,
            Post.discussion_id == discussion.id))
        post_subquery = post_subquery.group_by(intervals_table.c.interval_id)
        post_subquery = post_subquery.subquery()

        # The cumulative posters
        cumulative_posts_aliased = aliased(Post)
        cumulative_posts_subquery = discussion.db.query(intervals_table.c.interval_id,
                                                        func.count(distinct(cumulative_posts_aliased.id)).label('count_cumulative_posts'),
                                                        func.count(distinct(cumulative_posts_aliased.creator_id)).label('count_cumulative_post_authors')
                                                        # func.DB.DBA.BAG_AGG(cumulative_posts_aliased.id).label('cumulative_post_ids')
                                                        )
        cumulative_posts_subquery = cumulative_posts_subquery.outerjoin(cumulative_posts_aliased, and_(
            cumulative_posts_aliased.creation_date < intervals_table.c.interval_end,
            cumulative_posts_aliased.discussion_id == discussion.id))
        cumulative_posts_subquery = cumulative_posts_subquery.group_by(intervals_table.c.interval_id)
        cumulative_posts_subquery = cumulative_posts_subquery.subquery()

        # The top posters
        top_post_subquery = discussion.db.query(intervals_table.c.interval_id,
                                                func.count(distinct(Post.id)).label('count_top_posts'),
                                                func.count(distinct(Post.creator_id)).label('count_top_post_authors'),
                                                # func.DB.DBA.BAG_AGG(Post.creator_id).label('post_authors'),
                                                # func.DB.DBA.BAG_AGG(Post.id).label('post_ids'),
                                                )
        top_post_subquery = top_post_subquery.outerjoin(Post, and_(
            Post.creation_date >= intervals_table.c.interval_start,
            Post.creation_date < intervals_table.c.interval_end,
            Post.parent_id == None,
            Post.discussion_id == discussion.id))
        top_post_subquery = top_post_subquery.group_by(intervals_table.c.interval_id)
        top_post_subquery = top_post_subquery.subquery()

        # The cumulative posters
        cumulative_top_posts_aliased = aliased(Post)
        cumulative_top_posts_subquery = discussion.db.query(intervals_table.c.interval_id,
                                                            func.count(distinct(cumulative_top_posts_aliased.id)).label('count_cumulative_top_posts'),
                                                            func.count(distinct(cumulative_top_posts_aliased.creator_id)
                                                                       ).label('count_cumulative_top_post_authors')
                                                            # func.DB.DBA.BAG_AGG(cumulative_top_posts_aliased.id).label('cumulative_post_ids')
                                                            )
        cumulative_top_posts_subquery = cumulative_top_posts_subquery.outerjoin(cumulative_top_posts_aliased, and_(
            cumulative_top_posts_aliased.creation_date < intervals_table.c.interval_end,
            cumulative_top_posts_aliased.parent_id == None,
            cumulative_top_posts_aliased.discussion_id == discussion.id))
        cumulative_top_posts_subquery = cumulative_top_posts_subquery.group_by(intervals_table.c.interval_id)
        cumulative_top_posts_subquery = cumulative_top_posts_subquery.subquery()

        # The post viewers
        postViewers = aliased(ViewPost)
        viewedPosts = aliased(Post)
        post_viewers_subquery = discussion.db.query(intervals_table.c.interval_id,
                                                    func.count(distinct(postViewers.actor_id)).label('UNRELIABLE_count_post_viewers')
                                                    )
        post_viewers_subquery = post_viewers_subquery.outerjoin(postViewers, and_(
            postViewers.creation_date >= intervals_table.c.interval_start,
            postViewers.creation_date < intervals_table.c.interval_end)
        ).outerjoin(viewedPosts, and_(
            postViewers.post_id == viewedPosts.id,
            viewedPosts.discussion_id == discussion.id))
        post_viewers_subquery = post_viewers_subquery.group_by(intervals_table.c.interval_id)
        post_viewers_subquery = post_viewers_subquery.subquery()

        # The cumulative visitors
        cumulativeVisitorAgent = aliased(AgentStatusInDiscussion)
        cumulative_visitors_query = discussion.db.query(intervals_table.c.interval_id,
                                                        func.count(distinct(cumulativeVisitorAgent.id)).label('count_cumulative_logged_in_visitors'),
                                                        # func.DB.DBA.BAG_AGG(cumulativeVisitorAgent.id).label('first_time_visitors')
                                                        )
        cumulative_visitors_query = cumulative_visitors_query.outerjoin(cumulativeVisitorAgent, and_(
            cumulativeVisitorAgent.first_visit < intervals_table.c.interval_end,
            cumulativeVisitorAgent.discussion_id == discussion.id))
        cumulative_visitors_query = cumulative_visitors_query.group_by(intervals_table.c.interval_id)
        cumulative_visitors_subquery = cumulative_visitors_query.subquery()
        # query = cumulative_visitors_query

        # The members (can go up and down...)  Assumes that first_subscribed is available
        memberAgentStatus = aliased(AgentStatusInDiscussion)
        members_subquery = discussion.db.query(intervals_table.c.interval_id,
                                               func.count(memberAgentStatus.id).label('count_approximate_members')
                                               )
        members_subquery = members_subquery.outerjoin(memberAgentStatus, ((memberAgentStatus.last_unsubscribed >= intervals_table.c.interval_end) | (memberAgentStatus.last_unsubscribed.is_(
            None))) & ((memberAgentStatus.first_subscribed < intervals_table.c.interval_end) | (memberAgentStatus.first_subscribed.is_(None))) & (memberAgentStatus.discussion_id == discussion.id))
        members_subquery = members_subquery.group_by(intervals_table.c.interval_id)
        members_subquery = members_subquery.subquery()

        subscribersAgentStatus = aliased(AgentStatusInDiscussion)
        subscribers_query = discussion.db.query(intervals_table.c.interval_id,
                                                func.sum(
                                                    case([
                                                        (subscribersAgentStatus.last_visit == None, 0),
                                                        (and_(subscribersAgentStatus.last_visit < intervals_table.c.interval_end,
                                                              subscribersAgentStatus.last_visit >= intervals_table.c.interval_start), 1)
                                                    ], else_=0)
                                                ).label('retention_count_last_visit_in_period'),
                                                func.sum(
                                                    case([
                                                        (subscribersAgentStatus.first_visit == None, 0),
                                                        (and_(subscribersAgentStatus.first_visit < intervals_table.c.interval_end,
                                                              subscribersAgentStatus.first_visit >= intervals_table.c.interval_start), 1)
                                                    ], else_=0)
                                                ).label('recruitment_count_first_visit_in_period'),
                                                func.sum(
                                                    case([
                                                        (subscribersAgentStatus.first_subscribed == None, 0),
                                                        (and_(subscribersAgentStatus.first_subscribed < intervals_table.c.interval_end,
                                                              subscribersAgentStatus.first_subscribed >= intervals_table.c.interval_start), 1)
                                                    ], else_=0)
                                                ).label('recruitment_count_first_subscribed_in_period'),
                                                func.sum(
                                                    case([
                                                        (subscribersAgentStatus.last_unsubscribed == None, 0),
                                                        (and_(subscribersAgentStatus.last_unsubscribed < intervals_table.c.interval_end,
                                                              subscribersAgentStatus.last_unsubscribed >= intervals_table.c.interval_start), 1)
                                                    ], else_=0)
                                                ).label('retention_count_last_unsubscribed_in_period'),
                                                )
        subscribers_query = subscribers_query.outerjoin(subscribersAgentStatus, subscribersAgentStatus.discussion_id == discussion.id)
        subscribers_query = subscribers_query.group_by(intervals_table.c.interval_id)
        subscribers_subquery = subscribers_query.subquery()
        #query = subscribers_query

        # The votes
        votes_aliased = aliased(AbstractIdeaVote)
        votes_subquery = discussion.db.query(intervals_table.c.interval_id,
                                             func.count(distinct(votes_aliased.id)).label('count_votes'),
                                             func.count(distinct(votes_aliased.voter_id)).label('count_voters'),
                                             )
        votes_subquery = votes_subquery.outerjoin(Idea, Idea.discussion_id == discussion.id)
        votes_subquery = votes_subquery.outerjoin(votes_aliased, and_(
            votes_aliased.vote_date >= intervals_table.c.interval_start,
            votes_aliased.vote_date < intervals_table.c.interval_end,
            votes_aliased.idea_id == Idea.id))
        votes_subquery = votes_subquery.group_by(intervals_table.c.interval_id)
        votes_subquery = votes_subquery.subquery()

        # The cumulative posters
        cumulative_votes_aliased = aliased(AbstractIdeaVote)
        cumulative_votes_subquery = discussion.db.query(intervals_table.c.interval_id,
                                                        func.count(cumulative_votes_aliased.id).label('count_cumulative_votes'),
                                                        func.count(distinct(cumulative_votes_aliased.voter_id)).label('count_cumulative_voters')
                                                        )
        cumulative_votes_subquery = cumulative_votes_subquery.outerjoin(Idea, Idea.discussion_id == discussion.id)
        cumulative_votes_subquery = cumulative_votes_subquery.outerjoin(cumulative_votes_aliased, and_(
            cumulative_votes_aliased.vote_date < intervals_table.c.interval_end,
            cumulative_votes_aliased.idea_id == Idea.id))
        cumulative_votes_subquery = cumulative_votes_subquery.group_by(intervals_table.c.interval_id)
        cumulative_votes_subquery = cumulative_votes_subquery.subquery()

        content = with_polymorphic(
            Content, [], Content.__table__,
            aliased=False, flat=True)

        # The actions
        actions_on_post = discussion.db.query(
            intervals_table.c.interval_id.label('interval_id'), ActionOnPost.actor_id.label('actor_id'))
        actions_on_post = actions_on_post.outerjoin(content, content.discussion_id == discussion.id)
        actions_on_post = actions_on_post.outerjoin(ActionOnPost, and_(
            ActionOnPost.post_id == content.id,
            or_(and_(
                ActionOnPost.creation_date >= intervals_table.c.interval_start,
                ActionOnPost.creation_date < intervals_table.c.interval_end),
                and_(
                    ActionOnPost.tombstone_date >= intervals_table.c.interval_start,
                    ActionOnPost.tombstone_date < intervals_table.c.interval_end))))

        actions_on_idea = discussion.db.query(
            intervals_table.c.interval_id.label('interval_id'), ActionOnIdea.actor_id.label('actor_id'))
        actions_on_idea = actions_on_idea.outerjoin(Idea, Idea.discussion_id == discussion.id)
        actions_on_idea = actions_on_idea.outerjoin(ActionOnIdea, and_(
            ActionOnIdea.idea_id == Idea.id,
            or_(and_(
                ActionOnIdea.creation_date >= intervals_table.c.interval_start,
                ActionOnIdea.creation_date < intervals_table.c.interval_end),
                and_(
                    ActionOnIdea.tombstone_date >= intervals_table.c.interval_start,
                    ActionOnIdea.tombstone_date < intervals_table.c.interval_end))))

        posts = discussion.db.query(
            intervals_table.c.interval_id.label('interval_id'),
            Post.creator_id.label('actor_id'))
        posts = posts.outerjoin(Post, and_(
            Post.discussion_id == discussion.id,
            Post.creation_date >= intervals_table.c.interval_start,
            Post.creation_date < intervals_table.c.interval_end))

        actions_union_subquery = actions_on_post.union(actions_on_idea, posts).subquery()
        actions_subquery = discussion.db.query(intervals_table.c.interval_id,
                                               func.count(distinct(actions_union_subquery.c.actor_id)).label('count_actors')
                                               ).outerjoin(actions_union_subquery, actions_union_subquery.c.interval_id == intervals_table.c.interval_id
                                                           ).group_by(intervals_table.c.interval_id).subquery()

        # The actions
        cumulative_actions_on_post = discussion.db.query(
            intervals_table.c.interval_id.label('interval_id'), ActionOnPost.actor_id.label('actor_id'))
        cumulative_actions_on_post = cumulative_actions_on_post.outerjoin(content, content.discussion_id == discussion.id)
        cumulative_actions_on_post = cumulative_actions_on_post.outerjoin(ActionOnPost, and_(
            ActionOnPost.post_id == content.id,
            or_(ActionOnPost.creation_date < intervals_table.c.interval_end,
                ActionOnPost.tombstone_date < intervals_table.c.interval_end)))

        cumulative_actions_on_idea = discussion.db.query(
            intervals_table.c.interval_id.label('interval_id'), ActionOnIdea.actor_id.label('actor_id'))
        cumulative_actions_on_idea = cumulative_actions_on_idea.outerjoin(Idea, Idea.discussion_id == discussion.id)
        cumulative_actions_on_idea = cumulative_actions_on_idea.outerjoin(ActionOnIdea, and_(
            ActionOnIdea.idea_id == Idea.id,
            or_(ActionOnIdea.creation_date < intervals_table.c.interval_end,
                ActionOnIdea.tombstone_date < intervals_table.c.interval_end)))

        posts = discussion.db.query(
            intervals_table.c.interval_id.label('interval_id'),
            Post.creator_id.label('actor_id'))
        posts = posts.outerjoin(Post, and_(
            Post.discussion_id == discussion.id,
            Post.creation_date < intervals_table.c.interval_end))

        cumulative_actions_union_subquery = cumulative_actions_on_post.union(cumulative_actions_on_idea, posts).subquery()
        cumulative_actions_subquery = discussion.db.query(intervals_table.c.interval_id,
                                                          func.count(distinct(cumulative_actions_union_subquery.c.actor_id)).label('count_cumulative_actors')
                                                          ).outerjoin(cumulative_actions_union_subquery, cumulative_actions_union_subquery.c.interval_id == intervals_table.c.interval_id
                                                                      ).group_by(intervals_table.c.interval_id).subquery()

        combined_query = discussion.db.query(intervals_table,
                                             post_subquery,
                                             cumulative_posts_subquery,
                                             top_post_subquery,
                                             cumulative_top_posts_subquery,
                                             post_viewers_subquery,
                                             cumulative_visitors_subquery,
                                             votes_subquery,
                                             cumulative_votes_subquery,
                                             members_subquery,
                                             actions_subquery,
                                             cumulative_actions_subquery,
                                             case([
                                                 (cumulative_posts_subquery.c.count_cumulative_post_authors == 0, None),
                                                 (cumulative_posts_subquery.c.count_cumulative_post_authors != 0, (cast(post_subquery.c.count_post_authors,
                                                                                                                        Float) / cast(cumulative_posts_subquery.c.count_cumulative_post_authors, Float)))
                                             ]).label('fraction_cumulative_authors_who_posted_in_period'),
                                             case([
                                                 (cumulative_visitors_subquery.c.count_cumulative_logged_in_visitors == 0, None),
                                                 (cumulative_visitors_subquery.c.count_cumulative_logged_in_visitors != 0, (cast(
                                                     post_subquery.c.count_post_authors, Float) / cast(cumulative_visitors_subquery.c.count_cumulative_logged_in_visitors, Float)))
                                             ]).label('fraction_cumulative_logged_in_visitors_who_posted_in_period'),
                                             subscribers_subquery,
                                             )
        combined_query = combined_query.join(post_subquery, post_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(cumulative_posts_subquery, cumulative_posts_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(top_post_subquery, top_post_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(cumulative_top_posts_subquery, cumulative_top_posts_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(post_viewers_subquery, post_viewers_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(cumulative_visitors_subquery, cumulative_visitors_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(members_subquery, members_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(subscribers_subquery, subscribers_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(votes_subquery, votes_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(cumulative_votes_subquery, cumulative_votes_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(actions_subquery, actions_subquery.c.interval_id == intervals_table.c.interval_id)
        combined_query = combined_query.join(cumulative_actions_subquery, cumulative_actions_subquery.c.interval_id == intervals_table.c.interval_id)

        query = combined_query
        query = query.order_by(intervals_table.c.interval_id)
        results = query.all()

        intervals_table.drop(bind=bind)

    if format == JSON_MIMETYPE:
            # json default
        return Response(json.dumps(results, cls=DateJSONEncoder),
                        content_type='application/json')

    fieldnames = [
        "interval_id",
        "interval_start",
        "interval_end",
        "count_posts",
        "count_cumulative_posts",
        "count_top_posts",
        "count_cumulative_top_posts",
        "count_post_authors",
        "count_cumulative_post_authors",
        "fraction_cumulative_authors_who_posted_in_period",

        "count_votes",
        "count_cumulative_votes",
        "count_voters",
        "count_cumulative_voters",
        "count_actors",
        "count_cumulative_actors",

        "count_approximate_members",
        "count_first_time_logged_in_visitors",
        "count_cumulative_logged_in_visitors",
        "fraction_cumulative_logged_in_visitors_who_posted_in_period",
        "recruitment_count_first_visit_in_period",
        "recruitment_count_first_subscribed_in_period",
        "retention_count_last_visit_in_period",
        "retention_count_last_unsubscribed_in_period",
        "UNRELIABLE_count_post_viewers",
    ]
    # otherwise assume csv
    return csv_response([r._asdict() for r in results], format, fieldnames)


@view_config(context=InstanceContext, name="extract_csv_taxonomy",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_DISC_STATS)
def extract_taxonomy_csv(request):
    import assembl.models as m
    discussion = request.context._instance
    db = discussion.db
    extracts = db.query(m.Extract
        ).filter(m.Extract.discussion_id == discussion.id
        ).options(
            joinedload(m.Extract.tags_associations),
            joinedload_all('content.body.entries'),
        ).order_by(m.Extract.id).all()

    extract_list = []
    user_prefs = LanguagePreferenceCollection.getCurrent()
    fieldnames = ["Thematic", "Message", "Content harvested",  "Content locale",
                  "Original message", "Original locale", "Qualify by nature",
                  "Qualify by action", "Message owner full name", "Message owner username",
                  "Published on", "Harvester full name", "Harvester username",
                  "Harvested on", "Nugget", "State"]
    len_tags = max([len(extract.tags) for extract in extracts])
    tags_names = ["Tag{}".format(index + 1) for index in range(len_tags)]
    fieldnames.extend(tags_names)
    user_info_by_id = {}
    for extract in extracts:
        if extract.idea_id:
            thematic = db.query(m.Idea).get(extract.idea_id)
            if thematic:
                if thematic.title:
                    thematic = thematic.title.best_lang(user_prefs).value
                else:
                    thematic = no_thematic_associated
            else:
                thematic = no_thematic_associated
        else:
            thematic = no_thematic_associated
        if extract.locale_id:
            extract_locale = extract.locale.code
        else:
            extract_locale = "no extract locale"
        content = db.query(m.Post).get(extract.content_id)
        if content:
            if content.body:
                original_message = content.body.first_original().value
                original_locale = content.body.first_original().locale.code
                message = content.body.best_lang(user_prefs).value
            else:
                message = "no message"
        else:
            message = "no message"
        if not message:
            message = "no message"

        if thematic == no_thematic_associated:
            idea_ids = m.Idea.get_idea_ids_showing_post(content.id)
            for thematic_id in reversed(idea_ids):
                thematic_title = db.query(m.Idea).get(thematic_id).title
                if thematic_title:
                    thematic = thematic_title.best_lang(user_prefs).value
                    break
        if extract.body:
            content_harvested = extract.body
        else:
            content_harvested = "no content harvested"
        if extract.extract_nature:
            qualify_by_nature = extract.extract_nature.name
        else:
            qualify_by_nature = " "
        if extract.extract_action:
            qualify_by_action = extract.extract_action.name
        else:
            qualify_by_action = " "
        published_on = unicode(content.creation_date.replace(microsecond=0))
        harvested_on = unicode(extract.creation_date.replace(microsecond=0))
        nugget = "Yes" if extract.important else "No"
        state = getattr(extract, 'extract_state', ExtractStates.PUBLISHED.value)
        tags = sorted([t.value for t in extract.tags])

        user_info = user_info_by_id.get(content.creator_id, None)
        if user_info is None:
            owner_of_the_message = db.query(m.User).get(content.creator_id)
            user_info = {}
            user_info['fullname'] = (owner_of_the_message.real_name() or u"").encode('utf-8')
            user_info['username'] = (owner_of_the_message.username_p or u"").encode('utf-8')
            user_info_by_id[content.creator_id] = user_info

        message_full_name = user_info['fullname']
        message_username = user_info['username']

        user_info = user_info_by_id.get(extract.owner_id, None)
        if user_info is None:
            harvester = db.query(m.User).get(extract.owner_id)
            user_info = {}
            user_info['fullname'] = (harvester.real_name() or u"").encode('utf-8')
            user_info['username'] = (harvester.username_p or u"").encode('utf-8')
            user_info_by_id[extract.owner_id] = user_info

        harvester_full_name = user_info['fullname']
        harvester_username = user_info['username']

        extract_info = {
            "Thematic": thematic.encode('utf-8'),
            "Message": sanitize_text(message).encode('utf-8'),
            "Content harvested": content_harvested.encode('utf-8'),
            "Content locale": extract_locale.encode('utf-8'),
            "Original message": sanitize_text(original_message).encode('utf-8'),
            "Original locale": original_locale.encode('utf-8'),
            "Qualify by nature": qualify_by_nature.encode('utf-8'),
            "Qualify by action": qualify_by_action.encode('utf-8'),
            "Message owner full name": message_full_name,
            "Message owner username": message_username,
            "Published on": published_on.encode('utf-8'),
            "Harvester full name": harvester_full_name,
            "Harvester username": harvester_username,
            "Harvested on": harvested_on.encode('utf-8'),
            "Nugget": nugget.encode('utf-8'),
            "State": state.encode('utf-8')
        }
        len_tags = len(tags)
        extract_info.update(
            {tag_name: tags[index].encode('utf-8') if len_tags > index else "" for index, tag_name in enumerate(tags_names)})
        extract_list.append(extract_info)

    return csv_response(extract_list, CSV_MIMETYPE, fieldnames, content_disposition='attachment; filename="extract_taxonomies.csv"')


@view_config(context=InstanceContext, name="multi-module-export",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_DISC_STATS)
def multi_module_csv_export(request):
    as_buffer = asbool(request.GET.get('as_buffer', False))
    results = {sheet_name: None for sheet_name in sheet_names}
    fieldnames = {sheet_name: None for sheet_name in sheet_names}
    fieldnames['export_phase'], results['export_phase'] = phase_csv_export(request)
    fieldnames['export_module_survey'], results['export_module_survey'] = survey_csv_export(request)
    fieldnames['export_module_thread'], results['export_module_thread'] = thread_csv_export(request)
    fieldnames['export_module_multicolumns'], results['export_module_multicolumns'] = multicolumn_csv_export(request)
    fieldnames['vote_users_data'], results['vote_users_data'] = voters_csv_export(request)
    fieldnames['export_module_bright_mirror'], results['export_module_bright_mirror'] = bright_mirror_csv_export(request)
    fieldnames['export_module_vote'], results['export_module_vote'] = global_votes_csv_export(request)
    return csv_response_multiple_sheets(results, fieldnames, as_buffer=as_buffer)


@view_config(context=InstanceContext, name="users-export",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_DISC_STATS)
def users_csv_export(request):
    import assembl.models as m
    from collections import defaultdict
    discussion = request.context._instance
    db = discussion.db
    user_prefs = LanguagePreferenceCollection.getCurrent()

    ID = u"id".encode('utf-8')
    NAME = u"Nom prénom".encode('utf-8')
    EMAIL = u"Mail".encode('utf-8')
    USERNAME = u"Pseudo".encode('utf-8')
    CREATION_DATE = u"Date de création du compte".encode('utf-8')
    FIRST_VISIT = u"Date de la première connexion".encode('utf-8')
    LAST_VISIT = u"Date de la dernière connexion".encode('utf-8')
    TOP_POSTS = u"Nombre de Top posts".encode('utf-8')
    TOP_POST_REPLIES = u"Nombre de réponses reçues aux Top posts".encode('utf-8')
    POSTS = u"Nombre de posts".encode('utf-8')
    POSTS_REPLIES = u"Nombre de réponses reçues aux posts".encode('utf-8')
    TOTAL_POSTS = u"Nombre total de posts".encode('utf-8')
    AGREE_GIVEN = u'''Nombre de mentions "D'accord" donné'''.encode('utf-8')
    DISAGREE_GIVEN = u'''Nombre de mentions "Pas d'accord" donné'''.encode('utf-8')
    DONT_UNDERSTAND_GIVEN = u'''Nombre de mentions "Pas tout compris" donné'''.encode('utf-8')
    MORE_INFO_GIVEN = u'''Nombre de mentions "SVP + d'infos" donné'''.encode('utf-8')
    AGREE_RECEIVED = u'''Nombre de mentions "D'accord" reçu'''.encode('utf-8')
    DISAGREE_RECEIVED = u'''Nombre de mentions "Pas d'accord" reçu'''.encode('utf-8')
    DONT_UNDERSTAND_RECEIVED = u'''Nombre de mentions "Pas tout compris" reçu'''.encode('utf-8')
    MORE_INFO_RECEIVED = u'''Nombre de mentions "SVP + d'infos" reçu'''.encode('utf-8')
    IDEAS = u"Liste des idées dans lesquelles il a contribué".encode('utf-8')

    custom_fields = db.query(m.AbstractConfigurableField
                             ).filter(m.AbstractConfigurableField.discussion_id == discussion.id,
                      m.AbstractConfigurableField.identifier == 'CUSTOM',
             ).order_by(m.AbstractConfigurableField.order
             ).all()

    custom_fields_display = [(cf.id, cf.title.best_lang(user_prefs).value.encode('utf-8')) for cf in custom_fields]

    fieldnames = [NAME, EMAIL, USERNAME] \
                 + [cf[1] for cf in custom_fields_display] \
                 + [CREATION_DATE, FIRST_VISIT, LAST_VISIT, TOP_POSTS, TOP_POST_REPLIES, POSTS,
        POSTS_REPLIES, TOTAL_POSTS, AGREE_GIVEN, DISAGREE_GIVEN, DONT_UNDERSTAND_GIVEN, MORE_INFO_GIVEN, AGREE_RECEIVED,
        DISAGREE_RECEIVED, DONT_UNDERSTAND_RECEIVED, MORE_INFO_RECEIVED, IDEAS]
    LIKE_SENTIMENT = 'sentiment:like'
    DISLIKE_SENTIMENT = 'sentiment:disagree'
    DONT_UNDERSTAND_SENTIMENT = 'sentiment:dont_understand'
    MORE_INFO_SENTIMENT = 'sentiment:more_info'


    start, end, interval = get_time_series_timing(request)
    has_anon = asbool(request.GET.get('anon', False))

    posts = db.query(m.Post
        ).filter(m.Post.discussion_id == discussion.id
        ).filter(m.Post.creation_date > start
        ).filter(m.Post.creation_date < end
        ).order_by(m.Post.id).all()

    sentiment_counts = db.query(m.SentimentOfPost.post_id, m.SentimentOfPost.actor_id, m.SentimentOfPost.type).filter(
        m.SentimentOfPost.post_id.in_([post.id for post in posts]),
        m.SentimentOfPost.tombstone_condition()
    ).all()

    # Sort sentiments for better access depending of post/user
    sentiments_given_by_user = defaultdict(lambda: defaultdict(int))
    sentiments_received_by_post = defaultdict(lambda: defaultdict(int))

    for sentiment in sentiment_counts:
        sentiments_received_by_post[sentiment[0]][sentiment[2]] += 1
        sentiments_given_by_user[sentiment[1]][sentiment[2]] += 1

    users = {}

    users_in_discussion = db.query(m.User
      ).join(m.AgentStatusInDiscussion, m.AgentStatusInDiscussion.profile_id == m.User.id
      ).filter(m.AgentStatusInDiscussion.discussion_id == discussion.id
      ).order_by(m.User.id
      ).all()

    users_in_discussion_ids = [u.id for u in users_in_discussion]
    profile_fields_ids = [cf.id for cf in custom_fields]
    profile_field_values = db.query(m.ProfileField).filter(
        m.ProfileField.agent_profile_id.in_(users_in_discussion_ids),
        m.ProfileField.configurable_field_id.in_(profile_fields_ids)
    ).join(m.AbstractConfigurableField, m.ProfileField.configurable_field_id == m.AbstractConfigurableField.id
    ).order_by(m.ProfileField.agent_profile_id
    ).all()

    from itertools import groupby
    from operator import attrgetter

    user_profile_custom_values = {user_id: list(profile_fields) for user_id, profile_fields in groupby(profile_field_values, attrgetter('agent_profile_id'))}

    for user in users_in_discussion:
        # Sort out ppl who couldn't have connected during the required period
        if (user.first_visit and (user.creation_date > end or user.first_visit > end)) or (user.last_visit and user.last_visit < start):
            continue

        username = user.username_p.encode('utf-8') if user.username_p else ''
        email = user.get_preferred_email(has_anon).encode('utf-8') if user.get_preferred_email(has_anon) else ''
        users[user.id] = {
                NAME: user.name.encode('utf-8') if not has_anon else user.anonymous_name(),
                EMAIL: email,
                USERNAME: username if not has_anon else user.anonymous_username(),
                CREATION_DATE: user.creation_date.strftime('%Y-%m-%d %H:%M:%S') if user.creation_date else '',
                FIRST_VISIT: user.first_visit.strftime('%Y-%m-%d %H:%M:%S') if user.first_visit else '',
                LAST_VISIT: user.last_visit.strftime('%Y-%m-%d %H:%M:%S') if user.last_visit else '',
                AGREE_GIVEN: sentiments_given_by_user[user.id][LIKE_SENTIMENT],
                DISAGREE_GIVEN: sentiments_given_by_user[user.id][DISLIKE_SENTIMENT],
                DONT_UNDERSTAND_GIVEN: sentiments_given_by_user[user.id][DONT_UNDERSTAND_SENTIMENT],
                MORE_INFO_GIVEN: sentiments_given_by_user[user.id][MORE_INFO_SENTIMENT],
                AGREE_RECEIVED: 0,
                DISAGREE_RECEIVED: 0,
                DONT_UNDERSTAND_RECEIVED: 0,
                MORE_INFO_RECEIVED: 0,
                TOP_POSTS: 0,
                TOP_POST_REPLIES: 0,
                POSTS: 0,
                POSTS_REPLIES: 0,
                TOTAL_POSTS: 0,
                IDEAS: ''
            }
        custom_values = user_profile_custom_values.get(user.id, [])
        custom_values_dict = {cv.configurable_field_id: cv for cv in custom_values}
        for custom_field_id, custom_field_label in custom_fields_display:
            custom_value = custom_values_dict.get(custom_field_id, None)
            if not custom_value:
                formatted_value = str(custom_field_id) if custom_field_id is not None else ''
            else:
                field_type = custom_value.configurable_field.type
                if field_type == 'text_field':
                    formatted_value = custom_value.value_data['value']
                elif field_type == 'select_field':
                    value_id = custom_value.value_data['value'][0]
                    select_field_option = db.query(m.SelectFieldOption).get(get_primary_id(value_id))
                    formatted_value = select_field_option.label.best_lang(user_prefs).value
                else:
                    raise Exception('field type unhandled: {}'.format(field_type))

            users[user.id][custom_field_label] = formatted_value

        #Get SSO extra information
        for account in user.social_accounts:
            for key, value in account.extra_data.iteritems():
                if key not in fieldnames:
                    fieldnames.append(key)
                if key not in users[user.id].keys():
                    users[user.id][key] = value
                else:
                    users[user.id][key] += ', ' + value

    for post in posts:
        creator_id = post.creator.id
        if creator_id not in users.keys():
            continue

        if not post.id:
            # prevent assertion error just after
            continue

        descendants = post.get_descendants().all()

        if post.parent_id is None:
            users[creator_id][TOP_POSTS] += 1
            users[creator_id][TOP_POST_REPLIES] += len(descendants)
        else:
            users[creator_id][POSTS] += 1
            users[creator_id][POSTS_REPLIES] += len(descendants)

        users[creator_id][TOTAL_POSTS] += 1
        users[creator_id][AGREE_RECEIVED] += sentiments_received_by_post[post.id][LIKE_SENTIMENT]
        users[creator_id][DISAGREE_RECEIVED] += sentiments_received_by_post[post.id][DISLIKE_SENTIMENT]
        users[creator_id][DONT_UNDERSTAND_RECEIVED] += sentiments_received_by_post[post.id][DONT_UNDERSTAND_SENTIMENT]
        users[creator_id][MORE_INFO_RECEIVED] += sentiments_received_by_post[post.id][MORE_INFO_SENTIMENT]

        for idea in post.get_ideas():
            idea_title = idea.safe_title(user_prefs).encode('utf-8')
            if idea_title not in users[creator_id][IDEAS]:
                if len(users[creator_id][IDEAS]) == 0:
                    users[creator_id][IDEAS] = "%s(%d)" % (idea_title, idea.id)
                else:
                    users[creator_id][IDEAS] += ", %s(%d)" % (idea_title, idea.id)

    return csv_response(users.values(), CSV_MIMETYPE, fieldnames, content_disposition='attachment; filename="users.csv"')


def transform_fieldname(fn):
    if '_' in fn:
        return ' '.join(fn.split('_')).title()
    return fn


def csv_response(results, format, fieldnames=None, content_disposition=None, as_buffer=False):
    output = StringIO()
    if format == CSV_MIMETYPE:
        from csv import writer
        # include BOM for Excel to open the file in UTF-8 properly
        output.write(u'\ufeff'.encode('utf-8'))
        csv = writer(output, dialect='excel', delimiter=';')
        writerow = csv.writerow
        empty = ''
    elif format == XSLX_MIMETYPE:
        from zipfile import ZipFile, ZIP_DEFLATED
        from openpyxl.workbook import Workbook
        workbook = Workbook(True)
        archive = ZipFile(output, 'w', ZIP_DEFLATED, allowZip64=True)
        worksheet = workbook.create_sheet()
        writerow = worksheet.append
        empty = None

    if fieldnames:
        writerow([transform_fieldname(fn) for fn in fieldnames])
        for r in results:
            writerow([r.get(f, empty) for f in fieldnames])
    else:
        for r in results:
            writerow(r)

    if format == XSLX_MIMETYPE:
        from openpyxl.writer.excel import ExcelWriter
        writer = ExcelWriter(workbook, archive)
        writer.save('')
    output.seek(0)
    if not as_buffer:
        return Response(body_file=output, content_type=format, content_disposition=content_disposition)
    return output


def escapeit(r, data):
    """
    Return a suitable value for excel output from data. The trap here is sometimes we have a number, None,
    unicode or ascii, so we're trying not to break the export
    @param: r  A dict of rows which form the lookup of a value
    @param: data The value to find in the dict
    """
    import unicodedata
    try:
        something = r.get(data)
    except:
        print('r.get failed')
        return ''

    if something:
        if isinstance(something, basestring):
            # Clean control characters (their unicode category starts with C)
            try:
                return  "".join(ch for ch in remove_emoji(something.decode('utf-8')) if unicodedata.category(ch)[0]!="C")
            except:
                print('Error while decoding text:', something)
                return ''
        return something
    return ''


def remove_emoji(string):
    import re
    # Narrow UCS-2 build emoji removal
    emoji_pattern = re.compile(u'('
        u'\ud83c[\udf00-\udfff]|'
        u'\ud83d[\udc00-\ude4f\ude80-\udeff]|'
        u'[\u2600-\u26FF\u2700-\u27BF])+',
        re.UNICODE)
    return emoji_pattern.sub(' ', string)


def csv_response_multiple_sheets(results, fieldnames=None, content_disposition='attachment; filename=multimodule_excel_export.xlsx', as_buffer=False):
    """
    Return a multiple sheets excel file
    @param: results  A dict of lists. Each list contains dicts.
    @param: fieldnames A dict of lists. Each list contains a string.
    """
    output = StringIO()
    from zipfile import ZipFile, ZIP_DEFLATED
    from openpyxl.workbook import Workbook
    workbook = Workbook(True)
    empty = None
    archive = ZipFile(output, 'w', ZIP_DEFLATED, allowZip64=True)
    for sheet_name in sheet_names:
        workbook.create_sheet(sheet_name)

    for worksheet in workbook.worksheets:
        writerow = worksheet.append
        if fieldnames[worksheet.title] is not None:
            writerow([transform_fieldname(fn) for fn in fieldnames[worksheet.title]])
            if results[worksheet.title] is not None:
                for r in results[worksheet.title]:
                    try:
                        writerow([escapeit(r, f) for f in fieldnames[worksheet.title]])
                    except Exception:
                        print('Skipping row:', r)
        else:
            if results[worksheet.title] is not None:
                for r in results[worksheet.title]:
                    writerow(r)

    from openpyxl.writer.excel import ExcelWriter
    writer = ExcelWriter(workbook, archive)
    writer.save('')

    output.seek(0)
    if not as_buffer:
        return Response(body_file=output, content_type=XSLX_MIMETYPE, content_disposition=content_disposition)
    return output


@view_config(context=InstanceContext, name="contribution_count",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_DISC_STATS)
def get_contribution_count(request):
    start, end, interval = get_time_series_timing(request)
    format = get_format(request)
    discussion = request.context._instance
    results = []
    if interval < (end - start):
        while start < end:
            this_end = min(start + interval, end)
            results.append(dict(
                start=start, end=this_end,
                count=discussion.count_contributions_per_agent(
                    start, this_end)))
            start = this_end
    else:
        r = dict(count=discussion.count_contributions_per_agent(start, end))
        if not start:
            from assembl.models import Post
            (start,) = discussion.db.query(
                func.min(Post.creation_date)).filter_by(
                discussion_id=discussion.id).first()
        r["start"] = start
        if not end:
            end = datetime.now()
        r["end"] = end
        results.append(r)
    if format == JSON_MIMETYPE:
        # json default
        for v in results:
            v['count'] = {agent.display_name(): count
                          for (agent, count) in v['count']}
        return Response(json.dumps(results, cls=DateJSONEncoder),
                        content_type='application/json')

    total_count = defaultdict(int)
    agents = {}
    for v in results:
        as_dict = {}
        for (agent, count) in v['count']:
            total_count[agent.id] += count
            as_dict[agent.id] = count
            agents[agent.id] = agent
        v['count'] = as_dict
    count_list = total_count.items()
    count_list.sort(key=lambda (a, c): c, reverse=True)
    rows = []
    rows.append(['Start'] + [
        x['start'] for x in results] + ['Total'])
    rows.append(['End'] + [
        x['end'] for x in results] + [''])
    for agent_id, total_count in count_list:
        agent = agents[agent_id]
        agent_name = (
            agent.display_name() or agent.real_name() or
            agent.get_preferred_email())
        rows.append([agent_name.encode('utf-8')] + [
            x['count'].get(agent_id, '') for x in results] + [total_count])
    return csv_response(rows, format)


@view_config(context=InstanceContext, name="visit_count",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_DISC_STATS)
def get_visit_count(request):
    start, end, interval = get_time_series_timing(request)
    format = get_format(request)
    discussion = request.context._instance
    results = []
    if interval < (end - start):
        while start < end:
            this_end = min(start + interval, end)
            results.append(dict(
                start=start, end=this_end,
                readers=discussion.count_post_viewers(
                    start, this_end),
                first_visitors=discussion.count_new_visitors(
                    start, this_end)))
            start = this_end
    else:
        r = dict(
            readers=discussion.count_post_viewers(start, end),
            first_visitors=discussion.count_new_visitors(start, end))
        if not start:
            from assembl.models import AgentStatusInDiscussion
            (start,) = discussion.db.query(
                func.min(AgentStatusInDiscussion.first_visit)).filter_by(
                discussion_id=discussion.id).first()
        r["start"] = start
        if not end:
            end = datetime.now()
        r["end"] = end
        results.append(r)
    if format == JSON_MIMETYPE:
        # json default
        return Response(json.dumps(results, cls=DateJSONEncoder),
                        content_type='application/json')
    # otherwise assume csv
    fieldnames = ['start', 'end', 'first_visitors', 'readers']
    return csv_response(results, format, fieldnames)


@view_config(context=InstanceContext, name="visitors",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_DISC_STATS)
def get_visitors(request):
    discussion = request.context._instance
    user_prefs = LanguagePreferenceCollection.getCurrent()
    fieldnames = ["time", "name", "email"]
    extra_columns_info = (None if 'no_extra_columns' in request.GET else
                          load_social_columns_info(discussion, "en"))
    db = discussion.db
    from assembl import models as m
    # Adding configurable fields titles to the csv
    configurable_fields = db.query(m.AbstractConfigurableField).filter(m.AbstractConfigurableField.discussion_id == discussion.id).filter(
        m.AbstractConfigurableField.identifier == m.ConfigurableFieldIdentifiersEnum.CUSTOM.value).all()
    for configurable_field in configurable_fields:
        fieldnames.append((configurable_field.title.best_lang(user_prefs).value).encode("utf-8"))

    select_field_options = db.query(m.SelectFieldOption).all()
    select_field_options_dict = {sfd.id: sfd.label.best_lang(user_prefs).value for sfd in select_field_options}
    if extra_columns_info and not has_anon:
        # insert after email
        fieldnames.extend([name.encode('utf-8') for (name, path) in extra_columns_info])
        column_info_per_user = {}
        provider_id = get_provider_id_for_discussion(discussion)

    use_first = asbool(request.GET.get("first", False))
    attribute = "first_visit" if use_first else "last_visit"
    visitors = []
    for st in discussion.agent_status_in_discussion:
        if not getattr(st, attribute, None):
            continue
        profile_fields = db.query(m.AbstractConfigurableField, m.ProfileField).filter(m.ProfileField.discussion_id == discussion.id).join(m.ProfileField.configurable_field).filter(
            m.ProfileField.agent_profile_id == st.agent_profile.id).filter(m.AbstractConfigurableField.identifier == m.ConfigurableFieldIdentifiersEnum.CUSTOM.value).filter(m.ProfileField.configurable_field_id == m.AbstractConfigurableField.id).all()

        data = {"time": getattr(st, attribute),
                "name": (st.agent_profile.name or '').encode("utf-8"),
                "email": (st.agent_profile.get_preferred_email() or '').encode("utf-8")}

        for profile_field in profile_fields:
            if profile_field[1].value_data["value"] != None and profile_field[0].title != None:
                if type(profile_field[1].value_data["value"]) == list:
                    profile_field_value_id = profile_field[1].value_data["value"][0]
                    profile_field_value_id = int(Node.from_global_id(profile_field_value_id)[1])
                    profile_field_value = select_field_options_dict.get(profile_field_value_id)
                    if profile_field_value:
                        data.update({(profile_field[0].title.best_lang(user_prefs).value).encode(
                            "utf-8"): profile_field_value.encode("utf-8")})
                else:
                    data.update({(profile_field[0].title.best_lang(user_prefs).value).encode("utf-8"): (profile_field[1].value_data["value"]).encode("utf-8")})

        if extra_columns_info and not has_anon:
            extra_info = get_social_columns_from_user(
                st.agent_profile, extra_columns_info, provider_id)
            for num, (name, path) in enumerate(extra_columns_info):
                data[name] = extra_info[num]
        visitors.append(data)

    visitors.sort(key=lambda x: x['time'], reverse=True)
    return csv_response(visitors, CSV_MIMETYPE, fieldnames)


@view_config(context=InstanceContext, name="activity_alerts",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_DISC_STATS)
def get_activity_alerts(request):
    discussion = request.context._instance
    user_id = request.authenticated_userid or Everyone
    result = get_analytics_alerts(
        discussion, user_id,
        ["lurking_user", "inactive_user", "user_gone_inactive"],
        True)
    return Response(body=result, content_type='application/json')


@view_config(context=InstanceContext, name="interest_alerts",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_DISC_STATS)
def get_interest_alerts(request):
    discussion = request.context._instance
    user_id = request.authenticated_userid or Everyone
    result = get_analytics_alerts(
        discussion, user_id,
        ["interesting_to_me"],
        True)
    return Response(body=result, content_type='application/json')


@view_config(context=InstanceContext, name="test_results",
             ctx_instance_class=Discussion, request_method='POST',
             header=FORM_HEADER, permission=P_READ)
def test_results(request):
    mailer = get_mailer(request)
    config = get_config()
    message = Message(
        subject="test_results",
        sender=config.get('assembl.admin_email'),
        recipients=["maparent@acm.org"],
        body=json.dumps(request.POST.dict_of_lists()))
    mailer.send(message)
    return Response(body="Thank you!", content_type="text/text")


@view_config(context=InstanceContext, name="test_sentry",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_READ)
def test_sentry(request):
    raise RuntimeError("Let's test sentry")


@etalab_discussions.post(permission=P_SYSADMIN)
@view_config(context=ClassContext, ctx_class=Discussion,
             request_method='POST', header=JSON_HEADER, permission=P_SYSADMIN)
def post_discussion(request):
    from assembl.models import EmailAccount, User, LocalUserRole, Role, AbstractAgentAccount
    ctx = request.context
    json = request.json_body
    user_id = request.authenticated_userid or Everyone
    permissions = get_permissions(user_id, None)
    is_etalab_request = (request.matched_route and request.matched_route.name == 'etalab_discussions')
    if is_etalab_request:
        # The Etalab specification says that the API call representing the instance creation request must contain the following fields:
        # - requestIdentifier
        # - name: the title of the discussion (discussion.topic)
        # - slug
        # - adminName
        # - adminEmail
        default_view = 'etalab'
        # Fake an APIv2 context
        from ..traversal import Api2Context
        ctx = Api2Context(ctx)
        ctx = ClassContext(ctx, Discussion)
        json['topic'] = json.get('name', json.get('slug', ''))
    else:
        default_view = 'default'
    cls = ctx.get_class(json.get('@type', None))
    typename = cls.external_typename()
    # special case: find the user first.
    creator_email = json.get("adminEmail", None)
    db = Discussion.default_db
    if creator_email:
        account = db.query(AbstractAgentAccount).filter_by(
            email=creator_email, verified=True).first()
        if account:
            user = account.profile
        else:
            user = User(name=json.get("adminName", None), verified=True)
            account = EmailAccount(profile=user, email=creator_email, verified=True)
            db.add(user)
            db.flush()
        json['creator'] = user.uri()
    else:
        user = None
    try:
        instances = ctx.create_object(typename, json, user_id)
        discussion = instances[0]
        # Hackish. Discussion API? Generic post-init method?
        discussion.preferences.name = (
            'discussion_' + json.get('slug', str(discussion.id)))
        create_default_discussion_data(discussion)
        if user is not None:
            role = db.query(Role).filter_by(name=R_ADMINISTRATOR).first()
            local_role = LocalUserRole(discussion=discussion, user=user, role=role)
            instances.append(local_role)
        discussion.invoke_callbacks_after_creation()
    except ObjectNotUniqueError as e:
        raise HTTPConflict(e)
    except Exception as e:
        raise HTTPServerError(e)
    if instances:
        first = instances[0]
        db = first.db
        for instance in instances:
            db.add(instance)
        db.flush()
        view = request.GET.get('view', None) or default_view
        uri = "/".join((API_ETALAB_DISCUSSIONS_PREFIX, str(first.id))) if is_etalab_request else None
        return CreationResponse(first, user_id, permissions, view, uri=uri)


class defaultdict_of_dict(defaultdict):
    """A defaultdict of dicts."""

    def __init__(self):
        super(defaultdict_of_dict, self).__init__(dict)


@view_config(context=InstanceContext, name="visits_time_series_analytics",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_READ, renderer='json')  # TODO: What permission should this require? Do all debate initiators want this data to be accessible?
def get_visits_time_series_analytics(request):
    """
    Fetches visits analytics from bound piwik site.
    Optional parameters `start` and `end` are dates like "2017-11-21" (default dates are from discussion creation date to today as default).
    """

    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    discussion = request.context._instance
    try:
        data = discussion.get_visits_time_series_analytics(start, end)
        return data
    except ValueError as e:
        raise HTTPBadRequest(explanation=e)


@view_config(context=InstanceContext, name="participant_time_series_analytics",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_DISC_STATS)
def get_participant_time_series_analytics(request):
    start, end, interval = get_time_series_timing(request)
    data_descriptors = request.GET.getall("data")
    with_email = request.GET.get("email", None)
    discussion = request.context._instance
    user_id = request.authenticated_userid or Everyone
    permissions = get_permissions(user_id, discussion.id)
    if with_email is None:
        with_email = P_ADMIN_DISC in permissions
    else:
        with_email = asbool(with_email)
        if with_email and P_ADMIN_DISC not in permissions:
            raise HTTPUnauthorized("Cannot obtain email information")
    format = get_format(request)
    sort_key = request.GET.get('sort', 'domain' if with_email else 'name')
    results = []
    from assembl.models import (
        Post, AgentProfile, Idea,
        ActionOnPost, ActionOnIdea, Content,
        PublicationStates, AbstractAgentAccount, LikeSentimentOfPost,
        DisagreeSentimentOfPost, DontUnderstandSentimentOfPost,
        MoreInfoSentimentOfPost)

    default_data_descriptors = [
        "posts",
        "cumulative_posts",
        "top_posts",
        "cumulative_top_posts",
        "replies_received",
        "cumulative_replies_received",
        "active",
    ]
    sentiments = [
        ('liked', 'liking', LikeSentimentOfPost),
        ('disagreed', 'disagreeing', DisagreeSentimentOfPost),
        ('misunderstood', 'misunderstanding', DontUnderstandSentimentOfPost),
        ('info_requested', 'info_requesting', MoreInfoSentimentOfPost),
    ]
    for sentiment_in, sentiment_out, sentiment_class in sentiments:
        default_data_descriptors.extend([
            sentiment_in, 'cumulative_' + sentiment_in,
            sentiment_out, 'cumulative_' + sentiment_out])
    data_descriptors = data_descriptors or default_data_descriptors
    # Impose data_descriptors order
    data_descriptors = [s for s in default_data_descriptors if s in data_descriptors]
    if not data_descriptors:
        raise HTTPBadRequest("No valid data descriptor given")
    if sort_key and sort_key not in ('name', 'domain') and sort_key not in default_data_descriptors:
        raise HTTPBadRequest("Invalid sort column")
    if sort_key == 'domain' and P_ADMIN_DISC not in permissions:
        raise HTTPUnauthorized("Cannot obtain email information")

    with transaction.manager:
        bind = discussion.db.connection()
        metadata = MetaData(discussion.db.get_bind())  # make sure we are using the same connexion

        intervals_table = Table('temp_table_intervals_' + str(user_id), metadata,
                                Column('interval_id', Integer, primary_key=True),
                                Column('interval_start', DateTime, nullable=False),
                                Column('interval_end', DateTime, nullable=False),
                                prefixes=['TEMPORARY']
                                )
        # In case there is a leftover from a previous crash
        intervals_table.drop(bind=bind, checkfirst=True)
        intervals_table.create(bind=bind)
        interval_start = start
        intervals = []
        while interval_start < end:
            interval_end = min(interval_start + interval, end)
            intervals.append({'interval_start': interval_start, 'interval_end': interval_end})
            interval_start = interval_start + interval
        # pprint.pprint(intervals)
        discussion.db.execute(intervals_table.insert(), intervals)

        content = with_polymorphic(
            Content, [], Content.__table__,
            aliased=False, flat=True)
        # post = with_polymorphic(Post, [])

        query_components = []

        if 'posts' in data_descriptors:
            # The posters
            post_query = discussion.db.query(
                intervals_table.c.interval_id.label('interval_id_q'),
                AgentProfile.id.label('participant_id'),
                AgentProfile.name.label('participant'),
                literal('posts').label('key'),
                func.count(distinct(Post.id)).label('value'),
            )
            post_query = post_query.join(Post, and_(
                Post.creation_date >= intervals_table.c.interval_start,
                Post.creation_date < intervals_table.c.interval_end,
                Post.discussion_id == discussion.id))
            post_query = post_query.join(AgentProfile, Post.creator_id == AgentProfile.id)
            post_query = post_query.group_by(intervals_table.c.interval_id, AgentProfile.id)
            query_components.append(post_query)

        if 'cumulative_posts' in data_descriptors:
            # Cumulative posters
            cumulative_post_query = discussion.db.query(
                intervals_table.c.interval_id.label('interval_id_q'),
                AgentProfile.id.label('participant_id'),
                AgentProfile.name.label('participant'),
                literal('cumulative_posts').label('key'),
                func.count(distinct(Post.id)).label('value'),
            )
            cumulative_post_query = cumulative_post_query.join(Post, and_(
                Post.creation_date < intervals_table.c.interval_end,
                Post.publication_state == PublicationStates.PUBLISHED,
                Post.discussion_id == discussion.id))
            cumulative_post_query = cumulative_post_query.join(AgentProfile, Post.creator_id == AgentProfile.id)
            cumulative_post_query = cumulative_post_query.group_by(intervals_table.c.interval_id, AgentProfile.id)
            query_components.append(cumulative_post_query)

        if 'top_posts' in data_descriptors:
            # The posters
            top_post_query = discussion.db.query(
                intervals_table.c.interval_id.label('interval_id_q'),
                AgentProfile.id.label('participant_id'),
                AgentProfile.name.label('participant'),
                literal('top_posts').label('key'),
                func.count(distinct(Post.id)).label('value'),
            )
            top_post_query = top_post_query.join(Post, and_(
                Post.creation_date >= intervals_table.c.interval_start,
                Post.creation_date < intervals_table.c.interval_end,
                Post.parent_id == None,
                Post.discussion_id == discussion.id))
            top_post_query = top_post_query.join(
                AgentProfile, Post.creator_id == AgentProfile.id)
            top_post_query = top_post_query.group_by(
                intervals_table.c.interval_id, AgentProfile.id)
            query_components.append(top_post_query)

        if 'cumulative_top_posts' in data_descriptors:
            # Cumulative posters
            cumulative_top_post_query = discussion.db.query(
                intervals_table.c.interval_id.label('interval_id_q'),
                AgentProfile.id.label('participant_id'),
                AgentProfile.name.label('participant'),
                literal('cumulative_top_posts').label('key'),
                func.count(distinct(Post.id)).label('value'),
            )
            cumulative_top_post_query = cumulative_top_post_query.join(Post, and_(
                Post.creation_date < intervals_table.c.interval_end,
                Post.publication_state == PublicationStates.PUBLISHED,
                Post.parent_id == None,
                Post.discussion_id == discussion.id))
            cumulative_top_post_query = cumulative_top_post_query.join(
                AgentProfile, Post.creator_id == AgentProfile.id)
            cumulative_top_post_query = cumulative_top_post_query.group_by(
                intervals_table.c.interval_id, AgentProfile.id)
            query_components.append(cumulative_top_post_query)

        for sentiment_in, sentiment_out, sentiment_class in sentiments:
            c_sentiment_in = 'cumulative_' + sentiment_in
            c_sentiment_out = 'cumulative_' + sentiment_out
            if sentiment_out in data_descriptors:
                # The likes made
                query = discussion.db.query(
                    intervals_table.c.interval_id.label('interval_id_q'),
                    AgentProfile.id.label('participant_id'),
                    AgentProfile.name.label('participant'),
                    literal(sentiment_out).label('key'),
                    func.count(distinct(sentiment_class.id)).label('value'),
                )
                query = query.join(Post, Post.discussion_id == discussion.id)
                query = query.join(sentiment_class, and_(
                    sentiment_class.creation_date >= intervals_table.c.interval_start,
                    sentiment_class.creation_date < intervals_table.c.interval_end,
                    sentiment_class.post_id == Post.id))
                query = query.join(AgentProfile, sentiment_class.actor_id == AgentProfile.id)
                query = query.group_by(intervals_table.c.interval_id, AgentProfile.id)
                query_components.append(query)

            if c_sentiment_out in data_descriptors:
                # The cumulative active likes made
                query = discussion.db.query(
                    intervals_table.c.interval_id.label('interval_id_q'),
                    AgentProfile.id.label('participant_id'),
                    AgentProfile.name.label('participant'),
                    literal(c_sentiment_out).label('key'),
                    func.count(distinct(sentiment_class.id)).label('value'),
                )
                query = query.join(Post, Post.discussion_id == discussion.id)
                query = query.join(sentiment_class, and_(
                    sentiment_class.tombstone_date == None,
                    sentiment_class.creation_date < intervals_table.c.interval_end,
                    sentiment_class.post_id == Post.id))
                query = query.join(AgentProfile, sentiment_class.actor_id == AgentProfile.id)
                query = query.group_by(intervals_table.c.interval_id, AgentProfile.id)
                query_components.append(query)

            if sentiment_in in data_descriptors:
                # The likes received
                query = discussion.db.query(
                    intervals_table.c.interval_id.label('interval_id_q'),
                    AgentProfile.id.label('participant_id'),
                    AgentProfile.name.label('participant'),
                    literal(sentiment_in).label('key'),
                    func.count(distinct(sentiment_class.id)).label('value'),
                )
                query = query.join(Post, Post.discussion_id == discussion.id)
                query = query.join(sentiment_class, and_(
                    sentiment_class.creation_date >= intervals_table.c.interval_start,
                    sentiment_class.creation_date < intervals_table.c.interval_end,
                    sentiment_class.post_id == Post.id))
                query = query.join(AgentProfile, Post.creator_id == AgentProfile.id)
                query = query.group_by(intervals_table.c.interval_id, AgentProfile.id)
                query_components.append(query)

            if c_sentiment_in in data_descriptors:
                # The cumulative active likes received
                query = discussion.db.query(
                    intervals_table.c.interval_id.label('interval_id_q'),
                    AgentProfile.id.label('participant_id'),
                    AgentProfile.name.label('participant'),
                    literal(c_sentiment_in).label('key'),
                    func.count(distinct(sentiment_class.id)).label('value'),
                )
                query = query.outerjoin(Post, Post.discussion_id == discussion.id)
                query = query.outerjoin(sentiment_class, and_(
                    sentiment_class.tombstone_date == None,
                    sentiment_class.creation_date < intervals_table.c.interval_end,
                    sentiment_class.post_id == Post.id))
                query = query.outerjoin(AgentProfile, Post.creator_id == AgentProfile.id)
                query = query.group_by(intervals_table.c.interval_id, AgentProfile.id)
                query_components.append(query)

        if 'replies_received' in data_descriptors:
            # The posters
            reply_post = aliased(Post)
            original_post = aliased(Post)
            reply_post_query = discussion.db.query(
                intervals_table.c.interval_id.label('interval_id_q'),
                AgentProfile.id.label('participant_id'),
                AgentProfile.name.label('participant'),
                literal('replies_received').label('key'),
                func.count(distinct(reply_post.id)).label('value'),
            ).join(reply_post, and_(
                reply_post.creation_date >= intervals_table.c.interval_start,
                reply_post.creation_date < intervals_table.c.interval_end,
                reply_post.discussion_id == discussion.id)
            ).join(original_post, original_post.id == reply_post.parent_id
                   ).join(AgentProfile, original_post.creator_id == AgentProfile.id
                          ).group_by(intervals_table.c.interval_id, AgentProfile.id)
            query_components.append(reply_post_query)

        if 'cumulative_replies_received' in data_descriptors:
            # The posters
            reply_post = aliased(Post)
            original_post = aliased(Post)
            cumulative_reply_post_query = discussion.db.query(
                intervals_table.c.interval_id.label('interval_id_q'),
                AgentProfile.id.label('participant_id'),
                AgentProfile.name.label('participant'),
                literal('cumulative_replies_received').label('key'),
                func.count(distinct(reply_post.id)).label('value'),
            ).join(reply_post, and_(
                reply_post.creation_date < intervals_table.c.interval_end,
                reply_post.publication_state == PublicationStates.PUBLISHED,
                reply_post.discussion_id == discussion.id)
            ).join(original_post, and_(
                original_post.id == reply_post.parent_id,
                original_post.publication_state == PublicationStates.PUBLISHED)
            ).join(AgentProfile, original_post.creator_id == AgentProfile.id
                   ).group_by(intervals_table.c.interval_id, AgentProfile.id)
            query_components.append(cumulative_reply_post_query)

        if "active" in data_descriptors:
            actions_on_post = discussion.db.query(
                intervals_table.c.interval_id.label('interval_id'),
                ActionOnPost.actor_id.label('actor_id'),
                ActionOnPost.id.label('id'))
            actions_on_post = actions_on_post.join(content, content.discussion_id == discussion.id)
            actions_on_post = actions_on_post.join(ActionOnPost, and_(
                ActionOnPost.post_id == content.id,
                or_(and_(
                    ActionOnPost.creation_date >= intervals_table.c.interval_start,
                    ActionOnPost.creation_date < intervals_table.c.interval_end),
                    and_(
                        ActionOnPost.tombstone_date >= intervals_table.c.interval_start,
                        ActionOnPost.tombstone_date < intervals_table.c.interval_end))))

            actions_on_idea = discussion.db.query(
                intervals_table.c.interval_id.label('interval_id'),
                ActionOnIdea.actor_id.label('actor_id'),
                ActionOnIdea.id.label('id'))
            actions_on_idea = actions_on_idea.join(Idea, Idea.discussion_id == discussion.id)
            actions_on_idea = actions_on_idea.join(ActionOnIdea, and_(
                ActionOnIdea.idea_id == Idea.id,
                or_(and_(
                    ActionOnIdea.creation_date >= intervals_table.c.interval_start,
                    ActionOnIdea.creation_date < intervals_table.c.interval_end),
                    and_(
                        ActionOnIdea.tombstone_date >= intervals_table.c.interval_start,
                        ActionOnIdea.tombstone_date < intervals_table.c.interval_end))))

            posts = discussion.db.query(
                intervals_table.c.interval_id.label('interval_id'),
                Post.creator_id.label('actor_id'),
                Post.id.label('id'))
            posts = posts.join(Post, and_(
                Post.discussion_id == discussion.id,
                Post.creation_date >= intervals_table.c.interval_start,
                Post.creation_date < intervals_table.c.interval_end))

            actions_union_subquery = actions_on_post.union(actions_on_idea, posts).subquery()
            active_query = discussion.db.query(
                intervals_table.c.interval_id.label('interval_id_q'),
                AgentProfile.id.label('participant_id'),
                AgentProfile.name.label('participant'),
                literal('active').label('key'),
                cast(func.count(actions_union_subquery.c.id) > 0, Integer).label('value')
            ).join(actions_union_subquery, actions_union_subquery.c.interval_id == intervals_table.c.interval_id
                   ).join(AgentProfile, actions_union_subquery.c.actor_id == AgentProfile.id
                          ).group_by(intervals_table.c.interval_id, AgentProfile.id)
            query_components.append(active_query)

        combined_subquery = query_components.pop(0)
        if query_components:
            combined_subquery = combined_subquery.union(*query_components)
        combined_subquery = combined_subquery.subquery('combined')
        query = discussion.db.query(intervals_table, combined_subquery).outerjoin(
            combined_subquery, combined_subquery.c.interval_id_q == intervals_table.c.interval_id
        ).order_by(intervals_table.c.interval_id)
        results = query.all()
        intervals_table.drop(bind=bind)
        # pprint.pprint(results)
        # end of transaction

    if with_email:
        participant_ids = {row._asdict()['participant_id'] for row in results}
        # this is somewhat arbitrary...
        participant_emails = dict(
            discussion.db.query(AbstractAgentAccount.profile_id, AbstractAgentAccount.email
                                ).filter(AbstractAgentAccount.profile_id.in_(participant_ids),
                                         AbstractAgentAccount.verified == True,
                                         AbstractAgentAccount.email != None
                                         ).order_by(AbstractAgentAccount.preferred))

    if format == JSON_MIMETYPE:
        from assembl.lib.json import DateJSONEncoder
        combined = []
        interval_id = None
        interval_data = None
        interval_elements = ('interval_id', 'interval_start', 'interval_end')
        # We have fragmented interval+participant+key=>value.
        # Structure we're going for: List of intervals,
        # each data interval has list of combined participant info,
        # each in key=>value format.
        for element in results:
            element = element._asdict()
            if element['interval_id'] != interval_id:
                interval_data = {
                    k: element[k] for k in interval_elements
                }
                interval_data['data'] = interval_datalist = defaultdict(dict)
                combined.append(interval_data)
                interval_id = element['interval_id']
            participant_id = element['participant_id']
            if participant_id is not None:
                if element['value'] != 0:
                    data = interval_datalist[participant_id]
                    data[element['key']] = element['value']
                    data['participant'] = element['participant']
                    data['participant_id'] = participant_id
                    if with_email:
                        data['email'] = participant_emails.get(participant_id, '')
        for interval_data in combined:
            interval_data['data'] = interval_data['data'].values()
        return Response(json.dumps(combined, cls=DateJSONEncoder),
                        content_type=format)

    by_participant = defaultdict(defaultdict_of_dict)
    interval_ids = set()
    interval_starts = {}
    interval_ends = {}
    participant_names = {}
    email_column = int(with_email)

    for element in results:
        element = element._asdict()
        interval_id = element['interval_id']
        interval_ids.add(interval_id)
        interval_starts[interval_id] = element['interval_start']
        interval_ends[interval_id] = element['interval_end']
        pid = element['participant_id']
        value = element['value']
        if pid is not None and value != 0:
            participant_names[pid] = element['participant']
            key = element['key']
            by_participant[pid][interval_id][key] = value
    interval_ids = list(interval_ids)
    interval_ids.sort()
    num_cols = 2 + email_column + len(interval_ids) * len(data_descriptors)
    interval_starts = [interval_starts[id] for id in interval_ids]
    interval_ends = [interval_ends[id] for id in interval_ids]
    rows = []
    row = ['Participant id', 'Participant']
    if with_email:
        row.append('Email')
    for data_descriptor in data_descriptors:
        # TODO: i18n
        data_descriptor = ' '.join(data_descriptor.split('_')).title()
        row += [data_descriptor] * len(interval_ids)
    rows.append(row)
    empty_start = [''] * (1 + email_column)
    rows.append(empty_start + ['Interval id'] + interval_ids * len(data_descriptors))
    rows.append(empty_start + ['Interval start'] + interval_starts * len(data_descriptors))
    rows.append(empty_start + ['Interval end'] + interval_ends * len(data_descriptors))
    if sort_key == 'name':
        sorted_participants = [(name, id) for (id, name) in participant_names.iteritems()]
    elif sort_key == 'domain':
        sorted_participants = [(
            participant_emails.get(id, '').split('@')[-1],
            name, id) for (id, name) in participant_names.iteritems()]
    else:
        sorted_participants = [
            (-by_participant[id].get(interval_ids[-1], {}).get(sort_key, 0), id)
            for id in participant_names.iterkeys()]
    sorted_participants.sort()
    sorted_participants = [x[-1] for x in sorted_participants]
    for participant_id in sorted_participants:
        interval_data = by_participant[participant_id]
        row = [participant_id, participant_names[participant_id].encode('utf-8')]
        if with_email:
            email = participant_emails.get(participant_id, '') or ''
            row.append(email.encode('utf-8'))
        for data_descriptor in data_descriptors:
            row_part = [''] * len(interval_ids)
            for interval_id, data in interval_data.iteritems():
                row_part[interval_id - 1] = data.get(data_descriptor, '')
            row += row_part
        rows.append(row)

    return csv_response(rows, format)


def convert_to_utf8(rowdict):
    row = {}
    for key, value in rowdict.items():
        k = key.encode('utf-8') if isinstance(key, unicode) else key
        row[k] = value.encode('utf-8') if isinstance(value, unicode) else value

    return row


def get_entries_locale_original(lang_string):
    if lang_string is None:
        return {
            "entry": '',
            "original": '',
            "locale": ''
        }
    entries = lang_string.best_entries_in_request_with_originals()
    if len(entries) == 1:
        best = entries[0]
        original = entries[0]
    if len(entries) > 1:
        best = entries[0]
        original = entries[-1]
    locale = best.locale
    if locale is None:
        # zxx locale: best.locale_id=3 but best.locale is None
        locale = Locale.get(best.locale_id)
    locale_code = locale.code
    return {
        "entry": best.value,
        "original": original.value,
        "locale": locale_code
    }


IDEA_LEVEL_1 = u"Thématique niveau 1"
IDEA_LEVEL_2 = u"Thématique niveau 2"
IDEA_LEVEL_3 = u"Thématique niveau 3"
IDEA_LEVEL_4 = u"Thématique niveau 4"
IDEA_NAME = u"Nom de la thématique"
IDEA_PARENT = u"Thématique parent"
QUESTION_TITLE = u"Question"
POST_SUBJECT = u"Sujet"
POST_BODY = u"Message"
WORD_COUNT = u"Nombre de mots"
POST_CREATOR_NAME = u"Nom de l'auteur"
POST_CREATOR_USERNAME = u"Nom d'utilisateur de l'auteur"
POST_CREATOR_EMAIL = u"Adresse mail de l'auteur"
POST_CREATION_DATE = u"Date de publication"
POST_LIKE = u"J'aime"
POST_DISAGREE = u"J'aime pas"
POST_DONT_UNDERSTAND = u"Pas tout compris"
POST_MORE_INFO_PLEASE = u"SVP + d'infos"
SENTIMENT_ACTOR_NAME = u"Nom du votant"
SENTIMENT_ACTOR_EMAIL = u"Adresse mail du votant"
SENTIMENT_CREATION_DATE = u"Date du vote"
SHARE_COUNT = u"Nombre de share"
MESSAGE_URL = u"URL du message"
WATSON_SENTIMENT = u"Sentiment (analyse Watson)"
POST_CLASSIFIER = u"Nom de la colonne"

MESSAGE_INDENTATION = "Indentation du message"
TOP_POST_TITLE = "Titre du top post"
TOP_POST = "Top post"
TOP_POST_WORD_COUNT = "Nombre de mots top post"
POST_BODY_COUNT = u"Nombre de mots du post"
NUMBER_OF_ANSWERS = u"Nombre de réponses"

MESSAGE_COUNT = u"Nombre de message (débat)"
HARVESTING_COUNT = u"Nombre d'attrapages"
FICTION_URL = u"URL de la fiction"


def get_idea_parents_titles(idea, user_prefs):
    idea_title = idea.title.best_lang(user_prefs).value.encode("utf-8") if idea.title else ""
    ideas = []
    i = idea
    while i.parents and i.parents[0].sqla_type != u'root_idea' and not i.parents[0].hidden:
        i = i.parents[0]
        title = i.title.best_lang(user_prefs).value.encode("utf-8") if i.title else ""
        ideas.append(title)

    if len(ideas) == 0:
        return {
            IDEA_LEVEL_1: idea_title,
            IDEA_LEVEL_2: "",
            IDEA_LEVEL_3: "",
            IDEA_LEVEL_4: ""
        }
    if len(ideas) == 1:
        return {
            IDEA_LEVEL_1: ideas[0],
            IDEA_LEVEL_2: idea_title,
            IDEA_LEVEL_3: "",
            IDEA_LEVEL_4: ""
        }
    if len(ideas) == 2:
        return {
            IDEA_LEVEL_1: ideas[1],
            IDEA_LEVEL_2: ideas[0],
            IDEA_LEVEL_3: idea_title,
            IDEA_LEVEL_4: ""
        }
    if len(ideas) == 3:
        return {
            IDEA_LEVEL_1: ideas[2],
            IDEA_LEVEL_2: ideas[1],
            IDEA_LEVEL_3: ideas[0],
            IDEA_LEVEL_4: idea_title
        }


def phase_csv_export(request):
    """
    This is the first sheet of the multi-module export.
    The sheet is called export phase."""
    from assembl.models import Locale, Idea
    from assembl.models.auth import LanguagePreferenceCollection
    from assembl.utils import get_ideas_for_export
    start, end, interval = get_time_series_timing(request)
    has_lang = request.GET.get('lang', None)
    if has_lang:
        language = request.GET['lang']
        exists = Locale.get_id_of(language, create=False)
        if not exists:
            language = u'fr'
    else:
        language = u'fr'

    # This is required so that the langstring methods can operate using correct globals
    # on the request object
    LanguagePreferenceCollection.setCurrentFromLocale(language, req=request)
    user_prefs = LanguagePreferenceCollection.getCurrent()
    discussion = request.context._instance
    discussion_id = discussion.id
    Idea.prepare_counters(discussion_id, True)

    MODULE = u"Module"
    POSTED_MESSAGES_COUNT = u"Nombre de messages postés"
    DELETED_MESSAGES_COUNT = u"Nombre de messages supprimés"
    TOP_POST_COUNT = u"Nombre de Top post"
    NON_TOP_POST_COUNT = u"Nombre de messages (non top post)"
    LIKE = u"J'aime"
    DONT_LIKE = u"J'aime pas"
    DONT_UNDERSTAND = u"Pas tout compris"
    MORE_INFO = u"SVP + d'infos"
    THEMATIC_SHARE_COUNT = u"Nombre de share de la thématique"
    MESSAGE_SHARE_COUNT = u"Nombre de share de message"
    CONTRIBUTIONS_COUNT = u"Nombre de participations"
    CONTRIBUTORS_COUNT = u"Nombre de participants"
    fieldnames = [
        IDEA_LEVEL_1.encode('utf-8'),
        IDEA_LEVEL_2.encode('utf-8'),
        IDEA_LEVEL_3.encode('utf-8'),
        IDEA_LEVEL_4.encode('utf-8'),
        MODULE.encode('utf-8'),
        POSTED_MESSAGES_COUNT.encode('utf-8'),
        DELETED_MESSAGES_COUNT.encode('utf-8'),
        TOP_POST_COUNT.encode('utf-8'),
        NON_TOP_POST_COUNT.encode('utf-8'),
        LIKE.encode('utf-8'),
        DONT_LIKE.encode('utf-8'),
        DONT_UNDERSTAND.encode('utf-8'),
        MORE_INFO.encode('utf-8'),
        THEMATIC_SHARE_COUNT.encode('utf-8'),
        MESSAGE_SHARE_COUNT.encode('utf-8'),
        WATSON_SENTIMENT.encode('utf-8'), # TODO
        CONTRIBUTIONS_COUNT.encode('utf-8'),
        CONTRIBUTORS_COUNT.encode('utf-8')
    ]
    ideas = get_ideas_for_export(discussion, start=start, end=end)
    rows = []
    for idea in ideas:
        row = {}
        row.update(get_idea_parents_titles(idea, user_prefs))
        row[THEMATIC_SHARE_COUNT] = idea.share_count
        row[MODULE] = idea.message_view_override
        published_posts_query = get_published_posts(idea, start, end)
        row[POSTED_MESSAGES_COUNT] = published_posts_query.count()
        message_share_count_query = published_posts_query.with_entities(
            func.sum(Post.share_count)).order_by(None)
        row[MESSAGE_SHARE_COUNT] = message_share_count_query.first()[0] or 0
        top_key_words = idea.top_keywords()
        for index, key_word in enumerate(top_key_words):
            column_name = u"Mots clés {}".format(index + 1).encode('utf-8')
            if column_name not in fieldnames:
                fieldnames.append(column_name)
            row[column_name] = key_word.value.encode('utf-8')
        row[DELETED_MESSAGES_COUNT] = get_deleted_posts(idea, start, end).count()
        if idea.message_view_override == MessageView.thread.value:
            row[TOP_POST_COUNT] = get_published_top_posts(idea, start, end).count()
            row[NON_TOP_POST_COUNT] = row[POSTED_MESSAGES_COUNT] - row[TOP_POST_COUNT]
        else:
            row[TOP_POST_COUNT] = row[POSTED_MESSAGES_COUNT]
            row[NON_TOP_POST_COUNT] = 0
        row[LIKE] = idea.get_total_sentiments("like", start, end)
        row[DONT_LIKE] = idea.get_total_sentiments("disagree", start, end)
        row[DONT_UNDERSTAND] = idea.get_total_sentiments("dont_understand", start, end)
        row[MORE_INFO] = idea.get_total_sentiments("more_info", start, end)
        if idea.message_view_override == MessageView.voteSession.value:
            if not idea.vote_session:
                row[CONTRIBUTIONS_COUNT] = 0
                row[CONTRIBUTORS_COUNT] = 0
            else:
                row[CONTRIBUTIONS_COUNT] = idea.vote_session.get_num_votes(start, end)
                row[CONTRIBUTORS_COUNT] = idea.vote_session.get_voter_ids_query(start, end).count()
        else:
            row[CONTRIBUTORS_COUNT] = 0
            row[CONTRIBUTIONS_COUNT] = 0
        # To be implemented
        # row[WATSON_SENTIMENT] = idea.sentiments()
        rows.append(convert_to_utf8(row))
    return fieldnames, rows


def survey_csv_export(request):
    """CSV export for survey thematics."""
    from assembl.models import Locale, Idea
    start, end, interval = get_time_series_timing(request)
    has_anon = asbool(request.GET.get('anon', False))
    has_lang = request.GET.get('lang', None)
    if has_lang:
        language = request.GET['lang']
        exists = Locale.get_id_of(language, create=False)
        if not exists:
            language = u'fr'
    else:
        language = u'fr'

    # This is required so that the langstring methods can operate using correct globals
    # on the request object
    LanguagePreferenceCollection.setCurrentFromLocale(language, req=request)
    user_prefs = LanguagePreferenceCollection.getCurrent()
    discussion = request.context._instance
    discussion_id = discussion.id
    Idea.prepare_counters(discussion_id, True)

    POST_BODY = u"Réponse"
    fieldnames = [
        IDEA_LEVEL_1.encode('utf-8'),
        IDEA_LEVEL_2.encode('utf-8'),
        IDEA_LEVEL_3.encode('utf-8'),
        IDEA_LEVEL_4.encode('utf-8'),
        QUESTION_TITLE.encode('utf-8'),
        POST_BODY.encode('utf-8'),
        WORD_COUNT.encode('utf-8'),
        POST_CREATOR_NAME.encode('utf-8'),
        POST_CREATOR_USERNAME.encode('utf-8'),
        POST_CREATOR_EMAIL.encode('utf-8'),
        POST_CREATION_DATE.encode('utf-8'),
        POST_LIKE.encode('utf-8'),
        POST_DISAGREE.encode('utf-8'),
        SENTIMENT_ACTOR_NAME.encode('utf-8'),
        SENTIMENT_ACTOR_EMAIL.encode('utf-8'),
        SENTIMENT_CREATION_DATE.encode('utf-8'),
        SHARE_COUNT.encode('utf-8'),
        MESSAGE_URL.encode('utf-8'),
        WATSON_SENTIMENT.encode('utf-8')  # TODO
    ]

    extra_columns_info = (None if 'no_extra_columns' in request.GET else
                          load_social_columns_info(discussion, language))
    if extra_columns_info and not has_anon:
        # insert after email
        i = fieldnames.index(POST_CREATOR_EMAIL.encode('utf-8')) + 1
        fieldnames[i:i] = [name.encode('utf-8') for (name, path) in extra_columns_info]
        column_info_per_user = {}
        provider_id = get_provider_id_for_discussion(discussion)
        i = fieldnames.index(SENTIMENT_ACTOR_EMAIL.encode('utf-8')) + 1
        fieldnames[i:i] = ['sentiment ' + name.encode('utf-8') for (name, path) in extra_columns_info]

    thematics = get_survey_ideas(discussion, start, end)
    rows = []
    for thematic in thematics:
        for question in thematic.get_children():
            posts = get_published_posts(question, start, end)
            for post in posts:
                row = {}
                row.update(get_idea_parents_titles(thematic, user_prefs))
                row[QUESTION_TITLE] = get_entries_locale_original(question.title).get('entry')
                # To be implemented later
                # row[WATSON_SENTIMENT] = thematic.sentiments()
                if has_lang:
                    post.maybe_translate(target_locales=[language])

                body = get_entries_locale_original(post.body)
                row[POST_BODY] = sanitize_text(body.get('entry'))
                row[WORD_COUNT] = str(len(row[POST_BODY].split())) if row[POST_BODY] else "0"
                if not has_anon:
                    row[POST_CREATOR_NAME] = post.creator.real_name()
                    row[POST_CREATOR_USERNAME] = post.creator.username_p or ""
                else:
                    row[POST_CREATOR_NAME] = post.creator.anonymous_name()
                    row[POST_CREATOR_USERNAME] = post.creator.anonymous_username() or ""
                row[POST_CREATOR_EMAIL] = post.creator.get_preferred_email(anonymous=has_anon)
                row[POST_CREATION_DATE] = format_date(post.creation_date)
                row[MESSAGE_URL] = post.get_url()
                if extra_columns_info and not has_anon:
                    if post.creator_id not in column_info_per_user:
                        column_info_per_user[post.creator_id] = get_social_columns_from_user(
                            post.creator, extra_columns_info, provider_id)
                    extra_info = column_info_per_user[post.creator_id]
                    for num, (name, path) in enumerate(extra_columns_info):
                        row[name] = extra_info[num]

                row[SHARE_COUNT] = post.share_count

                for index, tag in enumerate(post.tags):
                    column_name = u"Mots clés {}".format(index + 1).encode('utf-8')
                    if column_name not in fieldnames:
                        fieldnames.append(column_name)
                    row[column_name] = tag.value.encode('utf-8')

                nlp_keywords = post.nlp_keywords()
                for index, key_word in enumerate(nlp_keywords):
                    column_name = u"Mots clés suggérés {}".format(index + 1).encode('utf-8')
                    if column_name not in fieldnames:
                        fieldnames.append(column_name)
                    row[column_name] = key_word.value.encode('utf-8')

                if post.sentiments:
                    row[POST_LIKE] = len([p for p in post.sentiments if p.name == 'like'])
                    row[POST_DISAGREE] = len([p for p in post.sentiments if p.name == 'disagree'])
                    for sentiment in post.sentiments:
                        if not has_anon:
                            row[SENTIMENT_ACTOR_NAME] = sentiment.actor.real_name()
                        else:
                            row[SENTIMENT_ACTOR_NAME] = sentiment.actor.anonymous_name()
                        row[SENTIMENT_ACTOR_EMAIL] = sentiment.actor.get_preferred_email(anonymous=has_anon)
                        row[SENTIMENT_CREATION_DATE] = format_date(sentiment.creation_date)
                        if extra_columns_info and not has_anon:
                            if sentiment.actor_id not in column_info_per_user:
                                column_info_per_user[sentiment.actor_id] = get_social_columns_from_user(
                                    sentiment.actor, extra_columns_info, provider_id)
                            extra_info = column_info_per_user[sentiment.actor_id]
                            for num, (name, path) in enumerate(extra_columns_info):
                                row['sentiment ' + name.encode('utf-8')] = extra_info[num]
                        rows.append(convert_to_utf8(row))
                else:
                    row[POST_LIKE] = 0
                    row[POST_DISAGREE] = 0
                    row[SENTIMENT_ACTOR_NAME] = u''
                    row[SENTIMENT_ACTOR_EMAIL] = u''
                    row[SENTIMENT_CREATION_DATE] = u''
                    rows.append(convert_to_utf8(row))
    return fieldnames, rows


def multicolumn_csv_export(request):
    """CSV export for the multicolumn sheet."""
    from assembl.models import Locale, Idea
    start, end, interval = get_time_series_timing(request)
    has_anon = asbool(request.GET.get('anon', False))
    has_lang = request.GET.get('lang', None)
    if has_lang:
        language = request.GET['lang']
        exists = Locale.get_id_of(language, create=False)
        if not exists:
            language = u'fr'
    else:
        language = u'fr'

    # This is required so that the langstring methods can operate using correct globals
    # on the request object
    LanguagePreferenceCollection.setCurrentFromLocale(language, req=request)
    user_prefs = LanguagePreferenceCollection.getCurrent()
    discussion = request.context._instance
    discussion_id = discussion.id
    Idea.prepare_counters(discussion_id, True)

    fieldnames = [
        IDEA_LEVEL_1.encode('utf-8'),
        IDEA_LEVEL_2.encode('utf-8'),
        IDEA_LEVEL_3.encode('utf-8'),
        IDEA_LEVEL_4.encode('utf-8'),
        POST_CLASSIFIER.encode('utf-8'),
        POST_BODY.encode('utf-8'),
        WORD_COUNT.encode('utf-8'),
        POST_CREATOR_NAME.encode('utf-8'),
        POST_CREATOR_USERNAME.encode('utf-8'),
        POST_CREATOR_EMAIL.encode('utf-8'),
        POST_CREATION_DATE.encode('utf-8'),
        POST_LIKE.encode('utf-8'),
        POST_DISAGREE.encode('utf-8'),
        POST_DONT_UNDERSTAND.encode('utf-8'),
        POST_MORE_INFO_PLEASE.encode('utf-8'),
        SENTIMENT_ACTOR_NAME.encode('utf-8'),
        SENTIMENT_ACTOR_EMAIL.encode('utf-8'),
        SENTIMENT_CREATION_DATE.encode('utf-8'),
        SHARE_COUNT.encode('utf-8'),
        MESSAGE_URL.encode('utf-8'),
        WATSON_SENTIMENT.encode('utf-8'),  # TODO
    ]
    extra_columns_info = (None if 'no_extra_columns' in request.GET else
                          load_social_columns_info(discussion, language))

    if extra_columns_info and not has_anon:
        # insert after email
        i = fieldnames.index(POST_CREATOR_EMAIL.encode('utf-8')) + 1
        fieldnames[i:i] = [name.encode('utf-8') for (name, path) in extra_columns_info]
        column_info_per_user = {}
        provider_id = get_provider_id_for_discussion(discussion)
        i = fieldnames.index(SENTIMENT_ACTOR_EMAIL.encode('utf-8')) + 1
        fieldnames[i:i] = ['sentiment ' + name.encode('utf-8') for (name, path) in extra_columns_info]

    ideas = get_multicolumns_ideas(discussion, start, end)
    rows = []
    for idea in ideas:
        row = {}
        top_key_words = idea.top_keywords()
        for index, key_word in enumerate(top_key_words):
            column_name = u"Mots clés {}".format(index + 1).encode('utf-8')
            if column_name not in fieldnames:
                fieldnames.append(column_name)
            row[column_name] = key_word.value.encode('utf-8')

        row.update(get_idea_parents_titles(idea, user_prefs))
        posts = get_published_posts(idea, start, end)
        # WATSON sentiment to be implemented later
        # row[WATSON_SENTIMENT] = idea.sentiments()
        for post in posts:
            if has_lang:
                post.maybe_translate(target_locales=[language])

            body = get_entries_locale_original(post.body)
            row[POST_BODY] = sanitize_text(body.get('entry'))
            row[WORD_COUNT] = str(len(row[POST_BODY].split())) if row[POST_BODY] else "0"
            idea_message_columns = idea.message_columns
            idea_message_column = [i for i in idea_message_columns if i.message_classifier == post.message_classifier]
            row[POST_CLASSIFIER] = idea_message_column[0].title.best_lang(user_prefs).value if idea_message_column else post.message_classifier
            if not has_anon:
                row[POST_CREATOR_NAME] = post.creator.real_name()
                row[POST_CREATOR_USERNAME] = post.creator.username_p or ""
            else:
                row[POST_CREATOR_NAME] = post.creator.anonymous_name()
                row[POST_CREATOR_USERNAME] = post.creator.anonymous_username() or ""
            row[POST_CREATOR_EMAIL] = post.creator.get_preferred_email(anonymous=has_anon)
            row[POST_CREATION_DATE] = format_date(post.creation_date)
            row[MESSAGE_URL] = post.get_url()
            if extra_columns_info and not has_anon:
                if post.creator_id not in column_info_per_user:
                    column_info_per_user[post.creator_id] = get_social_columns_from_user(
                        post.creator, extra_columns_info, provider_id)
                extra_info = column_info_per_user[post.creator_id]
                for num, (name, path) in enumerate(extra_columns_info):
                    row[name] = extra_info[num]

            row[SHARE_COUNT] = post.share_count
            if post.sentiments:
                row[POST_LIKE] = len([p for p in post.sentiments if p.name == 'like'])
                row[POST_DISAGREE] = len([p for p in post.sentiments if p.name == 'disagree'])
                row[POST_DONT_UNDERSTAND] = len([p for p in post.sentiments if p.name == 'dont_understand'])
                row[POST_MORE_INFO_PLEASE] = len([p for p in post.sentiments if p.name == 'more_info'])
                for sentiment in post.sentiments:
                    if not has_anon:
                        row[SENTIMENT_ACTOR_NAME] = sentiment.actor.real_name()
                    else:
                        row[SENTIMENT_ACTOR_NAME] = sentiment.actor.anonymous_name()
                    row[SENTIMENT_ACTOR_EMAIL] = sentiment.actor.get_preferred_email(anonymous=has_anon)
                    row[SENTIMENT_CREATION_DATE] = format_date(sentiment.creation_date)
                    if extra_columns_info and not has_anon:
                        if sentiment.actor_id not in column_info_per_user:
                            column_info_per_user[sentiment.actor_id] = get_social_columns_from_user(
                                sentiment.actor, extra_columns_info, provider_id)
                        extra_info = column_info_per_user[sentiment.actor_id]
                        for num, (name, path) in enumerate(extra_columns_info):
                            row['sentiment ' + name.encode('utf-8')] = extra_info[num]
                    rows.append(convert_to_utf8(row))
            else:
                row[POST_LIKE] = 0
                row[POST_DISAGREE] = 0
                row[POST_DONT_UNDERSTAND] = 0
                row[POST_MORE_INFO_PLEASE] = 0
                row[SENTIMENT_ACTOR_NAME] = u''
                row[SENTIMENT_ACTOR_EMAIL] = u''
                row[SENTIMENT_CREATION_DATE] = u''
                rows.append(convert_to_utf8(row))
    return fieldnames, rows


def get_latest_date(post):
    """From a post, get the latest creation_date of live descendants and self.
    """
    max_date = post.creation_date
    if len(post._children) == 0:
        if post.publication_state in deleted_publication_states:
            return None
        return max_date

    for p in post._children:
        date = get_latest_date(p)
        if date and date > max_date:
            max_date = date

    return max_date


# This is the equivalent of transformPosts in pages/idea.jsx to create the tree
def create_tree(posts):
    """Augment each post with _children and _indentation attributes.
    """
    posts_by_parent = {}
    posts_by_id = {p.id: p for p in posts}
    for p in posts:
        posts_by_parent.setdefault(p.parent_id, []).append(p)

    def get_post_indentation(post, ident=None):
        """Return post indentation with the form 1.2.1
        (first reply of the second post of the first top post)
        """
        if ident is None:
            ident = []

        if post.parent_id is None:
            try:
                pos = post._position
            except AttributeError:
                pos = 'x'
            ident[0:0] = [pos]
            return '.'.join([str(e) for e in ident])

        parent_post = posts_by_id[post.parent_id]
        parent_post_children = parent_post._children
        ident[0:0] = [parent_post_children.index(post) + 1]
        return get_post_indentation(parent_post, ident)

    def get_children(post_id):
        new_posts = []
        for p in posts_by_parent.get(post_id, ()):
            p._children = get_children(p.id)
            new_posts.append(p)
        new_posts.sort(key=lambda p: p.creation_date, reverse=True)
        return new_posts

    top_posts = []
    for p in posts_by_parent.get(None, ()):
        p._children = get_children(p.id)
        top_posts.append(p)

    top_posts = [p for p in top_posts
        if not (p.publication_state in deleted_publication_states and len(p._children) == 0)]

    top_posts.sort(key=lambda p: get_latest_date(p), reverse=True)
    for idx, p in enumerate(top_posts):
        p._position = idx + 1

    for p in posts:
        if p.publication_state == PublicationStates.PUBLISHED:
            p._indentation = get_post_indentation(p)
        else:
            p._indentation = 'x'

    return top_posts


def thread_csv_export(request):
    """CSV export for phase thread sheet"""
    from assembl.models import Locale, Idea
    start, end, interval = get_time_series_timing(request)
    has_anon = asbool(request.GET.get('anon', False))
    has_lang = request.GET.get('lang', None)
    if has_lang:
        language = request.GET['lang']
        exists = Locale.get_id_of(language, create=False)
        if not exists:
            language = u'fr'
    else:
        language = u'fr'

    # This is required so that the langstring methods can operate using correct globals
    # on the request object
    LanguagePreferenceCollection.setCurrentFromLocale(language, req=request)
    user_prefs = LanguagePreferenceCollection.getCurrent()
    discussion = request.context._instance
    discussion_id = discussion.id
    Idea.prepare_counters(discussion_id, True)

    fieldnames = [
        IDEA_LEVEL_1.encode('utf-8'),
        IDEA_LEVEL_2.encode('utf-8'),
        IDEA_LEVEL_3.encode('utf-8'),
        IDEA_LEVEL_4.encode('utf-8'),
        MESSAGE_INDENTATION.encode('utf-8'),
        TOP_POST_TITLE.encode('utf-8'),
        TOP_POST.encode('utf-8'),
        TOP_POST_WORD_COUNT.encode('utf-8'),
        POST_SUBJECT.encode('utf-8'),
        POST_BODY.encode('utf-8'),
        POST_BODY_COUNT.encode('utf-8'),
        NUMBER_OF_ANSWERS.encode('utf-8'),
        POST_CREATOR_NAME.encode('utf-8'),
        POST_CREATOR_USERNAME.encode('utf-8'),
        POST_CREATOR_EMAIL.encode('utf-8'),
        POST_CREATION_DATE.encode('utf-8'),
        POST_LIKE.encode('utf-8'),
        POST_DISAGREE.encode('utf-8'),
        POST_DONT_UNDERSTAND.encode('utf-8'),
        POST_MORE_INFO_PLEASE.encode('utf-8'),
        SENTIMENT_ACTOR_NAME.encode('utf-8'),
        SENTIMENT_ACTOR_EMAIL.encode('utf-8'),
        SENTIMENT_CREATION_DATE.encode('utf-8'),
        SHARE_COUNT.encode('utf-8'),
        MESSAGE_URL.encode('utf-8'),
        WATSON_SENTIMENT.encode('utf-8')  # TODO
    ]
    extra_columns_info = (None if 'no_extra_columns' in request.GET else
                          load_social_columns_info(discussion, language))

    if extra_columns_info and not has_anon:
        # insert after email
        i = fieldnames.index(POST_CREATOR_EMAIL.encode('utf-8')) + 1
        fieldnames[i:i] = [name.encode('utf-8') for (name, path) in extra_columns_info]
        column_info_per_user = {}
        provider_id = get_provider_id_for_discussion(discussion)
        i = fieldnames.index(SENTIMENT_ACTOR_EMAIL.encode('utf-8')) + 1
        fieldnames[i:i] = ['sentiment ' + name.encode('utf-8') for (name, path) in extra_columns_info]

    ideas = get_thread_ideas(discussion, start, end)
    rows = []
    for idea in ideas:
        children = idea.get_children()
        # We need to use get_posts without date filtering
        # instead of get_published_posts to create the tree.
        posts = get_posts(idea, None, None).all()
        # WATSON sentiment to be impemented later
        # row[WATSON_SENTIMENT] = idea.sentiments()
        create_tree(posts)  # this calculate p._indentation and p._children for each post
        for post in posts:
            row = {}
            row.update(get_idea_parents_titles(idea, user_prefs))

            if post.publication_state != PublicationStates.PUBLISHED:
                continue

            if post.creation_date < start or post.creation_date > end:
                continue

            if has_lang:
                post.maybe_translate(target_locales=[language])

            subject = get_entries_locale_original(post.subject)
            body = get_entries_locale_original(post.body)
            row[POST_SUBJECT] = subject.get('entry')
            top_post = post.get_top_post_in_thread()
            top_post_body = get_entries_locale_original(top_post.get_body())  # use get_body() instead of body, top post may be deleted
            top_post_title = get_entries_locale_original(top_post.get_subject())
            row[TOP_POST] = sanitize_text(top_post_body.get('entry'))
            row[TOP_POST_TITLE] = sanitize_text(top_post_title.get('entry'))
            row[TOP_POST_WORD_COUNT] = str(len(row[TOP_POST].split())) if row[TOP_POST] else "0"
            row[POST_BODY] = sanitize_text(body.get('entry'))
            row[POST_BODY_COUNT] = str(len(row[POST_BODY].split())) if row[POST_BODY] else "0"
            row[NUMBER_OF_ANSWERS] = len(post._children)
            row[MESSAGE_INDENTATION] = post._indentation
            row[MESSAGE_URL] = post.get_url()
            if not has_anon:
                row[POST_CREATOR_NAME] = post.creator.real_name()
                row[POST_CREATOR_USERNAME] = post.creator.username_p or ""
            else:
                row[POST_CREATOR_NAME] = post.creator.anonymous_name()
                row[POST_CREATOR_USERNAME] = post.creator.anonymous_username() or ""
            row[POST_CREATOR_EMAIL] = post.creator.get_preferred_email(anonymous=has_anon)
            row[POST_CREATION_DATE] = format_date(post.creation_date)
            if extra_columns_info and not has_anon:
                if post.creator_id not in column_info_per_user:
                    column_info_per_user[post.creator_id] = get_social_columns_from_user(
                        post.creator, extra_columns_info, provider_id)
                extra_info = column_info_per_user[post.creator_id]
                for num, (name, path) in enumerate(extra_columns_info):
                    row[name] = extra_info[num]

            row[SHARE_COUNT] = post.share_count

            for index, tag in enumerate(post.tags):
                column_name = u"Mots clés {}".format(index + 1).encode('utf-8')
                if column_name not in fieldnames:
                    fieldnames.append(column_name)
                row[column_name] = tag.value.encode('utf-8')

            nlp_keywords = post.nlp_keywords()
            for index, key_word in enumerate(nlp_keywords):
                column_name = u"Mots clés suggérés {}".format(index + 1).encode('utf-8')
                if column_name not in fieldnames:
                    fieldnames.append(column_name)
                row[column_name] = key_word.value.encode('utf-8')

            if post.sentiments:
                row[POST_LIKE] = len([p for p in post.sentiments if p.name == 'like'])
                row[POST_DISAGREE] = len([p for p in post.sentiments if p.name == 'disagree'])
                row[POST_DONT_UNDERSTAND] = len([p for p in post.sentiments if p.name == 'dont_understand'])
                row[POST_MORE_INFO_PLEASE] = len([p for p in post.sentiments if p.name == 'more_info'])
                for sentiment in post.sentiments:
                    if not has_anon:
                        row[SENTIMENT_ACTOR_NAME] = sentiment.actor.real_name()
                    else:
                        row[SENTIMENT_ACTOR_NAME] = sentiment.actor.anonymous_name()
                    row[SENTIMENT_ACTOR_EMAIL] = sentiment.actor.get_preferred_email(anonymous=has_anon)
                    row[SENTIMENT_CREATION_DATE] = format_date(sentiment.creation_date)
                    if extra_columns_info and not has_anon:
                        if sentiment.actor_id not in column_info_per_user:
                            column_info_per_user[sentiment.actor_id] = get_social_columns_from_user(
                                sentiment.actor, extra_columns_info, provider_id)
                        extra_info = column_info_per_user[sentiment.actor_id]
                        for num, (name, path) in enumerate(extra_columns_info):
                            row['sentiment ' + name.encode('utf-8')] = extra_info[num]
                    rows.append(convert_to_utf8(row))
            else:
                row[POST_LIKE] = 0
                row[POST_DISAGREE] = 0
                row[POST_DONT_UNDERSTAND] = 0
                row[POST_MORE_INFO_PLEASE] = 0
                row[SENTIMENT_ACTOR_NAME] = u''
                row[SENTIMENT_ACTOR_EMAIL] = u''
                row[SENTIMENT_CREATION_DATE] = u''
                rows.append(convert_to_utf8(row))
    return fieldnames, rows


def bright_mirror_csv_export(request):
    """CSV export for bright mirror sheet."""
    from assembl.models import Locale, Idea
    start, end, interval = get_time_series_timing(request)
    has_anon = asbool(request.GET.get('anon', False))
    has_lang = request.GET.get('lang', None)
    if has_lang:
        language = request.GET['lang']
        exists = Locale.get_id_of(language, create=False)
        if not exists:
            language = u'fr'
    else:
        language = u'fr'

    # This is required so that the langstring methods can operate using correct globals
    # on the request object
    LanguagePreferenceCollection.setCurrentFromLocale(language, req=request)
    user_prefs = LanguagePreferenceCollection.getCurrent()
    discussion = request.context._instance
    discussion_id = discussion.id
    Idea.prepare_counters(discussion_id, True)

    POST_SUBJECT = u"Titre de la fiction"
    POST_BODY = u"Fiction"
    fieldnames = [
        IDEA_LEVEL_1.encode('utf-8'),
        IDEA_LEVEL_2.encode('utf-8'),
        IDEA_LEVEL_3.encode('utf-8'),
        IDEA_LEVEL_4.encode('utf-8'),
        POST_SUBJECT.encode('utf-8'),
        POST_BODY.encode('utf-8'),
        WORD_COUNT.encode('utf-8'),
        POST_CREATOR_NAME.encode('utf-8'),
        POST_CREATOR_USERNAME.encode('utf-8'),
        POST_CREATOR_EMAIL.encode('utf-8'),
        POST_CREATION_DATE.encode('utf-8'),
        MESSAGE_COUNT.encode('utf-8'),
        HARVESTING_COUNT.encode('utf-8'),
        POST_LIKE.encode('utf-8'),
        POST_DISAGREE.encode('utf-8'),
        POST_DONT_UNDERSTAND.encode('utf-8'),
        POST_MORE_INFO_PLEASE.encode('utf-8'),
        SENTIMENT_ACTOR_NAME.encode('utf-8'),
        SENTIMENT_ACTOR_EMAIL.encode('utf-8'),
        SENTIMENT_CREATION_DATE.encode('utf-8'),
        SHARE_COUNT.encode('utf-8'),
        FICTION_URL.encode('utf-8'),
        WATSON_SENTIMENT.encode('utf-8')  # TODO
    ]
    extra_columns_info = (None if 'no_extra_columns' in request.GET else
                          load_social_columns_info(discussion, language))

    if extra_columns_info and not has_anon:
        # insert after email
        i = fieldnames.index(POST_CREATOR_EMAIL.encode('utf-8')) + 1
        fieldnames[i:i] = [name.encode('utf-8') for (name, path) in extra_columns_info]
        column_info_per_user = {}
        provider_id = get_provider_id_for_discussion(discussion)
        i = fieldnames.index(SENTIMENT_ACTOR_EMAIL.encode('utf-8')) + 1
        fieldnames[i:i] = ['sentiment ' + name.encode('utf-8') for (name, path) in extra_columns_info]

    ideas = get_bright_mirror_ideas(discussion, start, end)
    rows = []
    for idea in ideas:
        posts = get_published_top_posts(idea, start, end)  # we only care about fictions
        for post in posts:
            row = {}
            # WATSON sentiment to be impemented later
            # row[WATSON_SENTIMENT] = idea.sentiments()

            row.update(get_idea_parents_titles(idea, user_prefs))
            if has_lang:
                post.maybe_translate(target_locales=[language])

            subject = get_entries_locale_original(post.subject)
            body = get_entries_locale_original(post.body)
            row[POST_SUBJECT] = subject.get('entry')
            row[POST_BODY] = sanitize_text(body.get('entry'))
            row[WORD_COUNT] = str(len(row[POST_BODY].split())) if row[POST_BODY] else "0"
            if not has_anon:
                row[POST_CREATOR_NAME] = post.creator.real_name()
                row[POST_CREATOR_USERNAME] = post.creator.username_p or ""
            else:
                row[POST_CREATOR_NAME] = post.creator.anonymous_name()
                row[POST_CREATOR_USERNAME] = post.creator.anonymous_username() or ""
            row[POST_CREATOR_EMAIL] = post.creator.get_preferred_email(anonymous=has_anon)
            row[POST_CREATION_DATE] = format_date(post.creation_date)
            extracts = get_related_extracts(post)
            row[HARVESTING_COUNT] = extracts.count()
            row[MESSAGE_COUNT] = post.get_descendants().order_by(None).count()
            row[FICTION_URL] = post.get_url()
            if extra_columns_info and not has_anon:
                if post.creator_id not in column_info_per_user:
                    column_info_per_user[post.creator_id] = get_social_columns_from_user(
                        post.creator, extra_columns_info, provider_id)
                extra_info = column_info_per_user[post.creator_id]
                for num, (name, path) in enumerate(extra_columns_info):
                    row[name] = extra_info[num]

            row[SHARE_COUNT] = post.share_count

            for index, tag in enumerate(post.tags):
                column_name = u"Mots clés {}".format(index + 1).encode('utf-8')
                if column_name not in fieldnames:
                    fieldnames.append(column_name)
                row[column_name] = tag.value.encode('utf-8')

            nlp_keywords = post.nlp_keywords()
            for index, key_word in enumerate(nlp_keywords):
                column_name = u"Mots clés suggérés {}".format(index + 1).encode('utf-8')
                if column_name not in fieldnames:
                    fieldnames.append(column_name)
                row[column_name] = key_word.value.encode('utf-8')

            if post.sentiments:
                row[POST_LIKE] = len([p for p in post.sentiments if p.name == 'like'])
                row[POST_DISAGREE] = len([p for p in post.sentiments if p.name == 'disagree'])
                row[POST_DONT_UNDERSTAND] = len([p for p in post.sentiments if p.name == 'dont_understand'])
                row[POST_MORE_INFO_PLEASE] = len([p for p in post.sentiments if p.name == 'more_info'])
                for sentiment in post.sentiments:
                    if not has_anon:
                        row[SENTIMENT_ACTOR_NAME] = sentiment.actor.real_name()
                    else:
                        row[SENTIMENT_ACTOR_NAME] = sentiment.actor.anonymous_name()
                    row[SENTIMENT_ACTOR_EMAIL] = sentiment.actor.get_preferred_email(anonymous=has_anon)
                    row[SENTIMENT_CREATION_DATE] = format_date(sentiment.creation_date)
                    if extra_columns_info and not has_anon:
                        if sentiment.actor_id not in column_info_per_user:
                            column_info_per_user[sentiment.actor_id] = get_social_columns_from_user(
                                sentiment.actor, extra_columns_info, provider_id)
                        extra_info = column_info_per_user[sentiment.actor_id]
                        for num, (name, path) in enumerate(extra_columns_info):
                            row['sentiment ' + name.encode('utf-8')] = extra_info[num]
                    rows.append(convert_to_utf8(row))
            else:
                row[POST_LIKE] = 0
                row[POST_DISAGREE] = 0
                row[POST_DONT_UNDERSTAND] = 0
                row[POST_MORE_INFO_PLEASE] = 0
                row[SENTIMENT_ACTOR_NAME] = u''
                row[SENTIMENT_ACTOR_EMAIL] = u''
                row[SENTIMENT_CREATION_DATE] = u''
                rows.append(convert_to_utf8(row))
    return fieldnames, rows


def global_votes_csv_export(request):
    """CSV export for export_module_vote sheet."""
    from assembl.views.api2.votes import global_vote_results_csv
    from assembl.models import Locale, Idea
    start, end, interval = get_time_series_timing(request)
    has_lang = request.GET.get('lang', None)
    if has_lang:
        language = request.GET['lang']
        exists = Locale.get_id_of(language, create=False)
        if not exists:
            language = u'fr'
    else:
        language = u'fr'

    # This is required so that the langstring methods can operate using correct globals
    # on the request object
    LanguagePreferenceCollection.setCurrentFromLocale(language, req=request)
    user_prefs = LanguagePreferenceCollection.getCurrent()
    discussion = request.context._instance
    discussion_id = discussion.id
    Idea.prepare_counters(discussion_id, True)

    fieldnames = [
        IDEA_LEVEL_1.encode('utf-8'),
        IDEA_LEVEL_2.encode('utf-8'),
        IDEA_LEVEL_3.encode('utf-8'),
        IDEA_LEVEL_4.encode('utf-8')
    ]
    ideas = get_vote_session_ideas(discussion, start, end)
    votes_exports = {}
    for idea in ideas:
        if not idea.vote_session:
            continue
        votes_fieldnames, votes = global_vote_results_csv(idea.vote_session, request)
        votes_exports[idea.id] = votes
        for fieldname in votes_fieldnames:
            if fieldname not in fieldnames:
                fieldnames.append(fieldname)

    rows = []
    for idea in ideas:
        if not idea.vote_session:
            continue
        votes = votes_exports.get(idea.id)
        idea_levels = get_idea_parents_titles(idea, user_prefs)
        for vote_row in votes:
            row = {}
            row.update(idea_levels)
            row.update(vote_row)
            rows.append(convert_to_utf8(row))
    return fieldnames, rows


def voters_csv_export(request):
    """CSV export for vote_users_data sheet."""
    from assembl.views.api2.votes import extract_voters, VOTER_MAIL
    from assembl.models import Locale, Idea
    start, end, interval = get_time_series_timing(request)
    has_anon = asbool(request.GET.get('anon', False))
    has_lang = request.GET.get('lang', None)
    if has_lang:
        language = request.GET['lang']
        exists = Locale.get_id_of(language, create=False)
        if not exists:
            language = u'fr'
    else:
        language = u'fr'

    # This is required so that the langstring methods can operate using correct globals
    # on the request object
    LanguagePreferenceCollection.setCurrentFromLocale(language, req=request)
    user_prefs = LanguagePreferenceCollection.getCurrent()
    discussion = request.context._instance
    discussion_id = discussion.id
    Idea.prepare_counters(discussion_id, True)
    fieldnames = [
        IDEA_LEVEL_1.encode('utf-8'),
        IDEA_LEVEL_2.encode('utf-8'),
        IDEA_LEVEL_3.encode('utf-8'),
        IDEA_LEVEL_4.encode('utf-8')
    ]
    ideas = get_vote_session_ideas(discussion, start, end)
    votes_exports = {}
    for idea in ideas:
        if not idea.vote_session:
            continue
        votes_fieldnames, votes = extract_voters(idea.vote_session, request)
        votes_exports[idea.id] = votes
        for fieldname in votes_fieldnames:
            if fieldname not in fieldnames:
                fieldnames.append(fieldname)

    extra_columns_info = (None if 'no_extra_columns' in request.GET else
                          load_social_columns_info(discussion, language))
    if extra_columns_info and not has_anon:
        # insert after email
        i = fieldnames.index(VOTER_MAIL) + 1
        fieldnames[i:i] = [name.encode('utf-8') for (name, path) in extra_columns_info]
        column_info_per_user = {}
        provider_id = get_provider_id_for_discussion(discussion)

    rows = []
    for idea in ideas:
        if not idea.vote_session:
            continue
        votes = votes_exports.get(idea.id)
        idea_levels = get_idea_parents_titles(idea, user_prefs)
        for vote_row in votes:
            row = {}
            row.update(idea_levels)
            row.update(vote_row)
            rows.append(convert_to_utf8(row))
            if extra_columns_info and not has_anon:
                voter = vote_row['voter']
                if voter.id not in column_info_per_user:
                    column_info_per_user[voter.id] = get_social_columns_from_user(
                        voter, extra_columns_info, provider_id)
                extra_info = column_info_per_user[voter.id]
                for num, (name, path) in enumerate(extra_columns_info):
                    row[name] = extra_info[num]
    return fieldnames, rows


@view_config(context=InstanceContext, name="update_notification_subscriptions",
             ctx_instance_class=Discussion, request_method='GET',
             permission=P_ADMIN_DISC, renderer='json')
def update_notification_subscriptions(request):
    discussion = request.context._instance
    participants = discussion.all_participants
    for user in participants:
        user.get_notification_subscriptions(discussion.id)
    return {'status': 'Notification subscriptions have been updated.'}


discussions_slugs = Service(
    name='discussions_slugs',
    path='/discussions_slugs',
    description="List of existing Discussion slugs",
    renderer='json'
)


@discussions_slugs.get()
def get_discussions_slugs(request):
    user_id = request.authenticated_userid or Everyone
    discussions = discussions_with_access(user_id)
    return {'discussions': [{'id': discussion.id, 'slug': discussion.slug} for discussion in discussions]}


def includeme(config):
    # Make sure that the cornice view is registered
    pass
